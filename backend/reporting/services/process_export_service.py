"""
molding-optima 工艺参数导出服务

仅服务于 process app，不集成试模/模流相关的报告。
"""
from datetime import datetime
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from extensions.exceptions import BizException, ERROR_DATA_NOT_FOUND, ERROR_ILLEGAL_ARGUMENT
from reporting.utils import build_export_list_path


# 字段中文映射（精简版，按 molding-optima 的 ProcessParameter 实际字段）
PROCESS_PARAMETER_FIELDS = [
    ("parameter_code", "工艺参数编号"),
    ("sequence_index", "序列序号"),
    ("injection_stages", "注射段数"),
    ("IV0", "第一段注射速度"),
    ("IP0", "第一段注射压力"),
    ("IL0", "第一段注射位置"),
    ("IT", "注射时间"),
    ("IDT", "注射延时"),
    ("CT", "冷却时间"),
    ("MDT", "熔胶延时"),
    ("MEL", "熔胶终止位置"),
    ("barrel_temperature_stages", "料筒温度段数"),
    ("BT1", "第一段料筒温度"),
    ("BT2", "第二段料筒温度"),
    ("BT3", "第三段料筒温度"),
    ("BT4", "第四段料筒温度"),
    ("BT5", "第五段料筒温度"),
    ("updated_at", "更新时间"),
]


def export_process_parameters(
    tenant_slug: str,
    process_condition_ids: list,
    template: str = "default",
    include_snapshots: bool = False,
):
    """
    导出工艺参数到 Excel 文件。
    """
    # 延迟导入 process 模型，避免循环依赖
    from process.models import ProcessCondition, ProcessParameter

    if not process_condition_ids:
        raise BizException(ERROR_DATA_NOT_FOUND, "process_condition_ids must not be empty")

    conditions = ProcessCondition.objects.filter(
        id__in=process_condition_ids,
        is_deleted=False,
    ).prefetch_related("process_parameters")

    if not conditions.exists():
        raise BizException(ERROR_DATA_NOT_FOUND, "未找到任何匹配的工艺条件")

    template_name = f"process_parameters_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    full_path, relative_path = build_export_list_path(tenant_slug, template_name, "xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "工艺参数"

    # 表头
    headers = [zh for _, zh in PROCESS_PARAMETER_FIELDS]
    if include_snapshots:
        headers.append("工艺快照")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    row = 2
    for condition in conditions:
        params = list(condition.process_parameters.filter(is_deleted=False).order_by("sequence_index"))
        for param in params:
            for col_idx, (field, _) in enumerate(PROCESS_PARAMETER_FIELDS, start=1):
                value = getattr(param, field, None)
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row, column=col_idx, value=value)

            if include_snapshots:
                ws.cell(row=row, column=len(headers), value=condition.process_context_snapshot)

            row += 1

    # 列宽自适应
    for col_idx, header in enumerate(headers, start=1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else "AA"
        ws.column_dimensions[col_letter].width = 18

    wb.save(full_path)

    return {
        "file_path": full_path,
        "relative_path": relative_path,
        "count": row - 2,
        "template": template,
    }


def generate_process_report(
    tenant_slug: str,
    process_condition_id: int,
    format: str = "pdf",
    template: str = None,
):
    """
    生成工艺报告（PDF/DOCX）。molding-optima 阶段 1 暂仅返回占位结构，
    第二阶段将从 old/ 提炼报告生成逻辑。
    """
    from process.models import ProcessCondition

    condition = ProcessCondition.objects.filter(
        id=process_condition_id,
        is_deleted=False,
    ).first()

    if not condition:
        raise BizException(ERROR_DATA_NOT_FOUND, f"process condition not found: {process_condition_id}")

    template_name = f"process_report_{process_condition_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    full_path, relative_path = build_export_list_path(tenant_slug, template_name, format)

    # 阶段 1 占位：仅写入最小内容
    with open(full_path, "wb") as f:
        f.write(b"%PDF-1.4\n%Placeholder process report\n")

    return {
        "file_path": full_path,
        "relative_path": relative_path,
        "format": format,
        "template": template,
    }