# """
# 上传图片和视频服务
# """
import os
import struct
import hashlib
import filetype
import time
import decimal
import re
from openpyxl import load_workbook

from django.conf import settings
from django.db import transaction

from gis.common.exceptions import BizException
from hsmolding.exceptions import ERROR_NO_FILE, ERROR_ERROR_FILE_TYPE, ERROR_FILE_TYPE, ERROR_TEMPLATE
from hsmolding.const import SliceStatus, TEST_LIST
from hsmolding.services import moldflow_report_service, moldflow_service, project_service, machine_service, polymer_service
from hsmolding.models import Project, Product
from hsmolding.media_model import Upload
from hsmolding.utils.ppt.read_ppt import PPT

from mdprocess.services.process_index_service import add_process_index
from mdprocess.services.process_record_service import add_process_record
from mdprocess.utils.unit_convert import unit_conversion
from hsmolding.services.project_service import get_mold_dict_by_id
from hsmolding.services.machine_service import add_machine,update_machine
from hsmolding.services.polymer_service import add_polymer,update_polymer
from hsmolding.models import Machine, Polymer
from hsmolding.utils.code.get_code import detectCode
import time


global_language_code = None


def upload_file(request):
    # 从请求表单中获取文件对象
    file = request.FILES.get("file", None)

    if not file:  # 文件对象不存在， 返回400请求错误
        raise BizException(ERROR_NO_FILE)

    search_type = request.POST.get("search_type")

    if search_type == "media":
        if request.POST.get("slice_total"):
            return upload_slice(request)
        else:
            return upload_image(request)
    elif search_type == "polymer":
        return upload_polymer_file(request)
    elif search_type == "mold":
        return upload_mold_file(request)
    elif search_type == "mold_flow_ppt":
        return upoad_moldflow_ppt(request)


# 上传图片
def upload_image(request):
    company_id = request.user.get("company_id")
    file = request.FILES.get("file", None)
    search_id = request.POST.get("search_id", None)
    search_type = request.POST.get("search_type", None)

    with transaction.atomic():        
        md5 = calculate_md5(file)  # 计算图片文件md5
        file_type = get_file_extension(file)
        month = time.strftime("%Y-%m", time.localtime())

        make_folder(settings.FILE_STORAGE_PATH, "gsid_" + str(company_id) + "/media/" + month)

        file_name = file.name
        file_url = "gsid_" + str(company_id) + "/media/" + month + "/" + str(md5) + "." + file_type
        searched_file = get_file_by_url(file_url)  # 查看文件是否存在
        absolute_path = settings.FILE_STORAGE_PATH + file_url

        if not searched_file or searched_file and not os.path.exists(absolute_path):
            upload_dict: dict = save_upload(
                company_id=company_id,
                search_id=search_id,
                search_type=search_type,
                filename=file_name, 
                file_url=file_url,
                file_size=file.size,
                file_md5=md5, 
                file_type=file_type 
            )

            if not os.path.exists(absolute_path):
                save_file(absolute_path, file)

            return { "id": upload_dict.get("id"), "name": file_name, "url": file_url }
        return { "id": searched_file.id, "name": file_name, "url": file_url }


# 上传材料相关的文件
def upload_polymer_file(request):
    company_id = request.user.get("company_id")
    file = request.FILES.get("file", None)
    search_id = request.POST.get("search_id", None)
    search_type = request.POST.get("search_type", None)

    with transaction.atomic():        
        md5 = calculate_md5(file)  # 计算图片文件md5
        file_type = get_file_type(file.name)
        month = time.strftime("%Y-%m", time.localtime())

        make_folder(settings.FILE_STORAGE_PATH, "gsid_" + str(company_id) + "/polymer/" + month)

        file_name = file.name
        file_url = "gsid_" + str(company_id) + "/polymer/" + month + "/" + str(md5) + "." + file_type
        searched_file = get_file_by_url(file_url)  # 查看文件是否存在
        absolute_path = settings.FILE_STORAGE_PATH + file_url

        if not searched_file or searched_file and not os.path.exists(absolute_path):
            upload_dict: dict = save_upload(
                company_id=company_id,
                search_id=search_id,
                search_type=search_type,
                filename=file_name, 
                file_url=file_url,
                file_size=file.size,
                file_md5=md5, 
                file_type=file_type 
            )

            if not os.path.exists(absolute_path):
                save_file(absolute_path, file)

            return { "id": upload_dict.get("id"), "name": file_name, "url": file_url }
        return { "id": searched_file.id, "name": file_name, "url": file_url }


# 上传模具相关的文件
def upload_mold_file(request):
    company_id = request.user.get("company_id")
    file = request.FILES.get("file", None)
    search_id = request.POST.get("search_id", None)
    search_type = request.POST.get("search_type", None)

    with transaction.atomic():        
        md5 = calculate_md5(file)  # 计算图片文件md5
        file_type = get_file_type(file.name)
        month = time.strftime("%Y-%m", time.localtime())

        make_folder(settings.FILE_STORAGE_PATH, "gsid_" + str(company_id) + "/mold/" + month)

        file_name = file.name
        file_url = "gsid_" + str(company_id) + "/mold/" + month + "/" + str(md5) + "." + file_type
        searched_file = get_file_by_url(file_url)  # 查看文件是否存在
        absolute_path = settings.FILE_STORAGE_PATH + file_url

        if not searched_file or searched_file and not os.path.exists(absolute_path):
            upload_dict: dict = save_upload(
                company_id=company_id,
                search_id=search_id,
                search_type=search_type,
                filename=file_name, 
                file_url=file_url,
                file_size=file.size,
                file_md5=md5, 
                file_type=file_type 
            )

            if not os.path.exists(absolute_path):
                save_file(absolute_path, file)

            return { "id": upload_dict.get("id"), "name": file_name, "url": file_url }
        return { "id": searched_file.id, "name": file_name, "url": file_url }


# 上传大文件（如：视频）
def upload_slice(request):
    company_id = request.user.get("company_id")
    search_id = request.POST.get("search_id")
    search_type = request.POST.get("search_type")
    slice_total = request.POST.get("slice_total")
    slice_order = request.POST.get("slice_order")
    slice_md5 = request.POST.get("slice_md5")
    finally_md5 = request.POST.get("finally_md5")
    origin_filename = request.POST.get("origin_filename")
    file_type = origin_filename.split(".")[-1]
    month = time.strftime("%Y-%m", time.localtime())
    file = request.FILES.get("file", None)

    make_folder(settings.FILE_STORAGE_PATH, "gsid_" + str(company_id) + "/media/" + month)
    
    # 查看合并后的文件是否存在
    file_name = origin_filename
    file_url = "gsid_" + str(company_id) + "/media/" + month + "/" + str(finally_md5) + "." + file_type

    searched_file = get_file_by_url(file_url)
    if searched_file:
        return { "id": searched_file.id, "name": searched_file.filename, "url": searched_file.file_url }

    # 查看分片文件是否存在
    slice_file_url = "gsid_" + str(company_id) + "/media/slice/" + slice_md5
    searched_file = get_file_by_url(slice_file_url)
    absolute_path = settings.FILE_STORAGE_PATH + slice_file_url

    if not searched_file or searched_file and not os.path.exists(absolute_path):
        save_upload(
            company_id=company_id,
            search_id=search_id,
            search_type=search_type,
            filename=origin_filename,
            file_size=file.size,
            file_type=file_type,
            file_url=slice_file_url,
            file_md5=slice_md5,
            slice_status=SliceStatus.SLICE,
            slice_order=slice_order,
            slice_total=slice_total,
            finally_md5=finally_md5,
        )

        make_folder(settings.FILE_STORAGE_PATH, "gsid_" + str(company_id) + "/media/slice/")

        if not os.path.exists(absolute_path):
            save_file(absolute_path, file)

    if slice_order == slice_total:
        final_file_url = "gsid_" + str(company_id) + "/media/" + month + "/" + finally_md5 + "." + file_type
        finally_file_path = settings.FILE_STORAGE_PATH + final_file_url

        slice_files = get_slice_files(finally_md5)
        total_size = 0

        with open(finally_file_path, "wb") as outfile:
            for item in slice_files:
                total_size += item.file_size
                slice_file = settings.FILE_STORAGE_PATH + item.file_url
                with open(slice_file, "rb") as infile:
                    outfile.write(infile.read())
                os.remove(slice_file)
                
        # 保存数据库
        upload_dict: dict = save_upload(
            company_id=company_id, 
            search_id=search_id,
            search_type=search_type,
            filename=origin_filename, 
            file_url=final_file_url, 
            file_size=total_size, 
            file_md5=finally_md5, 
            file_type=file_type
        )

        return { "id": upload_dict.get("id"), "name": origin_filename, "url": final_file_url }


# 根据 search_id 跟 search_type 获取相关的文件
def get_list_of_file(
    search_id=None,
    search_type=None,
    file_url=None
):
    query = Upload.objects.all()
    if search_id:
        query = query.filter(search_id=search_id)
    if search_type:
        query = query.filter(search_type=search_type)
    if file_url:
        query = query.filter(file_url=file_url)
    file_dict_list = [ { "id": file.id, "name": file.filename, "url": file.file_url } for file in query ]
    return file_dict_list


def delete_file(file_id):
    file = Upload.objects.get(id=file_id)
    file.delete()


def make_folder(basic_path, folder):
    if not os.path.exists(basic_path + folder):
        os.makedirs(basic_path + folder)


def get_slice_files(finally_md5):
    return Upload.objects.filter(finally_md5=finally_md5).filter(slice_status=SliceStatus.SLICE).order_by("slice_order")


# 更新upload的表
def save_upload(
    company_id=None,
    filename=None,
    file_md5=None,
    file_type=None,
    file_size=None,
    file_url="",
    slice_status=0,
    slice_order=0,
    slice_total=0,
    finally_md5="",
    search_id=None,
    search_type=None
):
    # 检测通过 创建新的上传对象
    upload = Upload.objects.filter(file_url=file_url).first()
    if not upload:
        upload = Upload()

    upload.company_id = company_id
    upload.filename = filename
    upload.file_url = file_url
    upload.file_size = file_size
    upload.file_md5 = file_md5
    upload.file_type = file_type
    upload.slice_status = slice_status
    upload.slice_order = slice_order
    upload.slice_total = slice_total
    upload.finally_md5 = finally_md5
    upload.search_id = search_id
    upload.search_type = search_type

    try:
        upload.save()
        updated = Upload.objects.filter(file_url=file_url).first()
        return updated.to_dict()
    except Exception as e:
        raise Exception()
    


# 保存 文件到磁盘
def save_file(absolute_path, file):
    with open(absolute_path, "wb+") as f:
        # 分块写入
        for chunk in file.chunks():
            f.write(chunk)


# 计算文件的md5
def calculate_md5(file):
    md5_obj = hashlib.md5()
    for chunk in file.chunks():
        md5_obj.update(chunk)
    return md5_obj.hexdigest()


# 通过文件md5值获取模型对象的类方法
def get_file_by_md5(md5):
    return Upload.objects.filter(file_md5=md5).first()


# 查询文件是否存在
def get_file_by_url(file_url):
    return Upload.objects.filter(file_url=file_url).first()


# 检测文件类型
# 我们使用第三方的库filetype进行检测，而不是通过文件名进行判断
# pip install filetype 即可安装该库
def get_file_extension(file):
    raw_data = bytearray()
    for c in file.chunks():
        raw_data += c

    extension = filetype.guess_extension(raw_data)
    return extension


def get_file_type(file_name: str):
    if file_name:
        arr = file_name.split(".")
        file_type = arr[-1]
        return file_type


# 判断文件类型，上面的方法不支持excel的判断
def is_excel(file):
    binary_file = open(file, "rb")  # 二制字读取
    num_of_bytes = len("D0CF11E0")
    binary_file.seek(0)  # 每次读取都要回到文件头，不然会一直往后读取
    hbytes = struct.unpack_from("B" * num_of_bytes, binary_file.read(num_of_bytes))  # 一个 "B"表示一个字节
    if hbytes.hex() == "D0CF11E0":
        return True


# 从excel上传模具
def upload_mold(request):
    company_id = request.user.get("company_id")
    file = request.FILES.get("file", None)
    with transaction.atomic():
        md5 = calculate_md5(file)
        file_type = file.name.split(".")[-1]
        if "xls" not in file.name:
            return {"error_message":"该文件类型为"+file_type+",请使用模板EXCEL"}
        month = time.strftime("%Y%m", time.localtime())
        if not os.path.exists(settings.MOLD_PATH + str(company_id) + "/" + month):
            os.makedirs(settings.MOLD_PATH + str(company_id) + "/" + month)
        if "xls" in file.name:
            file_url = month + "/" + str(md5) + ".xlsx"

            absolute_path = settings.MOLD_PATH + str(company_id) + "/" + file_url
            if not os.path.exists(absolute_path):
                save_file(absolute_path, file)
            return absolute_path


def upoad_moldflow_ppt(request):
    project_id = request.POST.get("search_id", None)
    mold_flow_no = request.POST.get("mold_flow_no", None)
    # 从请求表单中获取文件对象
    file = request.FILES.get("file", None)
    if not file:  # 文件对象不存在， 返回400请求错误
        raise BizException(ERROR_NO_FILE)
    # 先判断文件格式是否合格
    file_type = get_file_type(file.name)
    if file_type != "ppt" and file_type != "pptx":
        raise BizException(ERROR_FILE_TYPE)
    else:
        absolute_path = settings.FILE_STORAGE_PATH + "gsid_"+str(request.user.get("company_id"))+"/moldflow_report/project_id_"+project_id+"/"
        if not os.path.exists(absolute_path):
            os.makedirs(absolute_path)
        model_path = absolute_path + file.name
        save_file(model_path, file)
        ppt_link = "gsid_"+str(request.user.get("company_id"))+"/moldflow_report/project_id_"+project_id+"/" + file.name

        # 从txt中获得分析结果list
        moldflow_dict = moldflow_report_service.get_moldflow_dict(project_id, mold_flow_no)
        if moldflow_dict:
            test_list = moldflow_report_service.get_moldflow_dict(project_id, mold_flow_no).get("test_list")
            # 解析ppt中的分析结果
            if test_list:
                ppt = PPT(filepath=model_path, test_list=test_list, company_id=request.user.get("company_id"), project_id=project_id)
            else:
                ppt = PPT(filepath=model_path, test_list=TEST_LIST, company_id=request.user.get("company_id"), project_id=project_id)
            ppt.main()
            moldflow_report = moldflow_report_service.add_moldflow({"ppt_link":ppt_link, "project_id": project_id, "mold_flow_no":mold_flow_no})
            return moldflow_report
        else:
            print("没有找到相应的模流分析文件")


# 上传moldflow日志
def upload_moldflow_log(request, mold_no, project_id, company_id):
    # 从请求表单中获取文件对象
    file = request.FILES.get("file", None)
    if not file:  # 文件对象不存在， 返回400请求错误
        raise BizException(ERROR_NO_FILE)
    # 先判断文件格式是否合格
    file_type = get_file_type(file.name)
    if file_type != "txt":
        raise BizException(ERROR_ERROR_FILE_TYPE)
    else:
        absolute_path = settings.FILE_STORAGE_PATH + "gsid_"+str(company_id)+"/moldflow/project_id_"+project_id+"/"
        if not os.path.exists(absolute_path):
            os.makedirs(absolute_path)
        model_path = absolute_path + "/" + mold_no + ".txt"
        save_file(model_path, file)

        params = moldflow_service.read_moldflow(model_path)
        if params and params!={} and not is_valid(params):
            doc_link = "gsid_"+str(company_id)+"/moldflow/" + mold_no + "/" + mold_no + ".txt"
            moldflow_service.add_moldflow({"monitor_item":params, "doc_link":doc_link, "project_id": project_id})
            return {"temp_moldflow":params, "doc_link":doc_link}
        else:
            raise BizException(ERROR_ERROR_FILE_TYPE)


def is_valid(params):
    isNone = True
    for key in params:
        if key not in ["inject_para", "holding_para"] and params.get(key):
            isNone = False
    return isNone


# 上传moldflow vbs宏脚本跑出来的日志
def upload_moldflow(request, mold_no, project_id, company_id):
    # 从请求表单中获取文件对象
    file = request.FILES.get("file", None)
    if not file:  # 文件对象不存在， 返回400请求错误
        raise BizException(ERROR_NO_FILE)
    # 先判断文件格式是否合格
    file_type = get_file_type(file.name)
    if file_type != "txt":
        raise BizException(ERROR_ERROR_FILE_TYPE)
    else:
        absolute_path = settings.FILE_STORAGE_PATH + "gsid_"+str(company_id)+"/moldflow_report/project_id_" + str(project_id)
        if not os.path.exists(absolute_path):
            os.makedirs(absolute_path)
        model_path = absolute_path + "/" + mold_no + ".txt"
        save_file(model_path, file)

        result = moldflow_report_service.get_mold_flow_data(model_path)
        params = result["mold_flow_data"]
        process_data = result["process_data"]
        setting_result = result["setting_result"]
        material_trademark = result["material_trademark"]
        material_abbreviation = result["material_abbreviation"]
        machine_manufacturer = result["machine_manufacturer"]
        machine_trademark = result["machine_trademark"]
        process_setting = result["process_setting"]
        start_ram_position = result["start_ram_position"]
        velocity_unit = setting_result["velocity_unit"]
        velocity_sec = setting_result["velocity_sec"]
        position_unit = setting_result["position_unit"]
        position_sec = setting_result["position_sec"]
        holding_control = setting_result["holding_control"]
        pressure_sec = setting_result["pressure_sec"]
        time_sec = setting_result["time_sec"]
        global_txt_path = result["global_txt_path"]
        result_tcode = result["result_tcode"]
        cooling_time = result["cooling_time"]
        if params and params!={} and not is_valid(params):
            doc_link = "gsid_"+str(company_id)+"/moldflow_report/project_id_" + str(project_id) + "/" + mold_no + ".txt"
            params["doc_link"] = doc_link
            params["project_id"] = project_id
            params["mold_flow_no"] = "M" + time.strftime("%Y%m%d%H%M%S", time.localtime())
            params["deleted"] = 0
            params["mold_no"] = mold_no
            moldflow_report_service.add_moldflow(params)

            # 获取编码方式
            global global_language_code
            global_language_code = detectCode(global_txt_path)

            # 从模流获得机器,如果机器数据库没有,则新增,获取ID
            moldflow_machine = None
            machine_info = {
                    "company_id":company_id,
                    "data_source":"模流", 
                    "trademark":machine_trademark,
                    "power_method":"液压机",
                    "manufacturer":machine_manufacturer,
                    "max_clamping_force":get_array(global_txt_path,tcode="10000-machine array:")[0] if get_array(global_txt_path,tcode="10000-machine array:") else None,
                    "pressure_unit": "MPa",
                "injectors_info":[{
                    "screw_diameter":get_array(global_txt_path,tcode="10008-machine array:")[0] if get_array(global_txt_path,tcode="10008-machine array:") else None,
                    "max_injection_stroke":get_array(global_txt_path,tcode="10007-machine array:")[0] if get_array(global_txt_path,tcode="10007-machine array:") else None,
                    "max_injection_rate":get_array(global_txt_path,tcode="10005-machine array:")[0] if get_array(global_txt_path,tcode="10005-machine array:") else None,
                    "intensification_ratio":get_array(global_txt_path,tcode="10004-machine array:")[0] if get_array(global_txt_path,tcode="10004-machine array:") else None,
                    "response_time":get_array(global_txt_path,tcode="10006-machine array:")[0] if get_array(global_txt_path,tcode="10006-machine array:") else None,
                    "max_set_injection_pressure":get_array(global_txt_path,tcode="10002-machine array:")[0] if get_array(global_txt_path,tcode="10002-machine array:") else None,
                    "max_injection_pressure":get_array(global_txt_path,tcode="10003-machine array:")[0] if get_array(global_txt_path,tcode="10003-machine array:") else None,
                    "max_set_holding_pressure":get_array(global_txt_path,tcode="10002-machine array:")[0] if get_array(global_txt_path,tcode="10002-machine array:") else None,
                    "max_holding_pressure":get_array(global_txt_path,tcode="10003-machine array:")[0] if get_array(global_txt_path,tcode="10003-machine array:") else None,
                    "max_set_metering_pressure":get_array(global_txt_path,tcode="10002-machine array:")[0] if get_array(global_txt_path,tcode="10002-machine array:") else None,
                    "max_metering_pressure":get_array(global_txt_path,tcode="10003-machine array:")[0] if get_array(global_txt_path,tcode="10003-machine array:") else None,
                    "max_set_decompression_pressure":get_array(global_txt_path,tcode="10002-machine array:")[0] if get_array(global_txt_path,tcode="10002-machine array:") else None,
                    "max_decompression_pressure":get_array(global_txt_path,tcode="10003-machine array:")[0] if get_array(global_txt_path,tcode="10003-machine array:") else None,
                    }]}
            machines = Machine.objects.filter(data_source="模流",trademark=machine_trademark,company_id=company_id).first()
            if not machines:
                moldflow_machine = add_machine(machine_info)
            else:
                moldflow_machine = update_machine(machines.id, machine_info)

            # 从模流获得材料,如果材料数据库没有,则新增,获取ID
            moldflow_polymer = None
            polymers = Polymer.objects.filter(data_source="模流",trademark=material_trademark, abbreviation=material_abbreviation,company_id=company_id).first()
            polymer_info = {
                    "company_id":company_id,
                    "data_source":"模流", 
                    "trademark":material_trademark,
                    "abbreviation":material_abbreviation,
                    "min_melt_temperature":get_array(global_txt_path,tcode="1800-material array:")[0] if get_array(global_txt_path,tcode="1800-material array:") else None,
                    "max_melt_temperature":get_array(global_txt_path,tcode="1800-material array:")[1] if get_array(global_txt_path,tcode="1800-material array:") else None,
                    "recommend_melt_temperature":get_array(global_txt_path,tcode="1801-material array:")[0] if get_array(global_txt_path,tcode="1801-material array:") else None,
                    "min_mold_temperature":get_array(global_txt_path,tcode="1808-material array:")[0] if get_array(global_txt_path,tcode="1808-material array:") else None,
                    "max_mold_temperature":get_array(global_txt_path,tcode="1808-material array:")[1] if get_array(global_txt_path,tcode="1808-material array:") else None,
                    "recommend_mold_temperature":get_array(global_txt_path,tcode="11108-material array:")[0] if get_array(global_txt_path,tcode="11108-material array:") else None,
                    "max_sheer_rate":get_array(global_txt_path,tcode="1806-material array:")[0] if get_array(global_txt_path,tcode="1806-material array:") else None,
                    "max_sheer_stress":get_array(global_txt_path,tcode="1804-material array:")[0] if get_array(global_txt_path,tcode="1804-material array:") else None,
                    "ejection_temperature":get_array(global_txt_path,tcode="1504-material array:")[0] if get_array(global_txt_path,tcode="1504-material array:") else None,
                    "degradation_temperature":get_array(global_txt_path,tcode="1550-material array:")[0] if get_array(global_txt_path,tcode="1550-material array:") else None,
                    "melt_density":get_array(global_txt_path,tcode="1000-material array:")[0] if get_array(global_txt_path,tcode="1000-material array:") else None,
                    "solid_density":get_array(global_txt_path,tcode="1001-material array:")[0] if get_array(global_txt_path,tcode="1001-material array:") else None,
                    }
            if not polymers:
                moldflow_polymer = add_polymer(polymer_info)
            else:
                moldflow_polymer = update_polymer(polymers.id, polymer_info)

            product = None
            # project = get_project_obj_by_id(project_id)
            project = get_mold_dict_by_id(project_id)
            product_infos = project.get("product_infos")
            if product_infos: 
                product = product_infos[0]
            # 第一步,保存到process_index
            process_index = {
                "company_id": company_id,
                "status": 2,
                "process_no": "P" + time.strftime("%Y%m%d%H%M%S", time.localtime()),
                "data_sources": "模流分析",
                "mold_trials_no": None,
                "mold_id": project.get("id"),
                "mold_no": mold_no,
                "cavity_num": project.get("cavity_num"),
                "runner_length": product.get("runner_length"),
                "runner_weight": product.get("runner_weight"),
                "gate_type": product.get("gate_type"),
                "gate_num": product.get("gate_num"),
                "gate_shape": product.get("gate_shape"),
                "gate_area": product.get("gate_area"),
                "gate_radius": product.get("gate_radius"),
                "gate_length": product.get("gate_length"),
                "gate_width": product.get("gate_width"),
                "inject_part": None,
                "product_no": project.get("product_no"),
                "product_category": project.get("product_category"),
                "product_type": project.get("product_type"),
                "product_name": project.get("product_name"),
                "product_total_weight": project.get("product_total_weight"),
                "product_ave_thickness": product.get("ave_thickness"),
                "product_max_thickness": product.get("max_thickness"),
                "product_max_length": product.get("flow_length"),
                "machine_id": moldflow_machine.get("id"),
                "machine_data_source": "模流",
                "machine_trademark": machine_trademark,
                "machine_serial_no": None,
                "polymer_id": moldflow_polymer.get("id"),
                "polymer_abbreviation": None,
                "polymer_trademark": material_trademark,
                "injection_stage": process_data.get("inject_stage"),
                "holding_stage": process_data.get("holding_stage"),
                "VP_switch_mode": None,
                "metering_stage": None,
                "decompressure_mode_befor": None,
                "decompressure_mode_after": None,
                "barrel_temperature_stage": None,
            }
            process = add_process_index(process_index)
            # 第二步,保存到process_record
            precondition = {
                "mold_id": project.get("id"),
                "data_sources": "模流分析",
                "mold_trials_no": None,
                "mold_no": mold_no,
                "cavity_num": project.get("cavity_num"),
                "runner_length": product.get("runner_length"),
                "runner_weight": product.get("runner_weight"),
                "gate_type": product.get("gate_type"),
                "gate_num": product.get("gate_num"),
                "gate_shape": product.get("gate_shape"),
                "gate_area": product.get("gate_area"),
                "gate_radius": product.get("gate_radius"),
                "gate_length": product.get("gate_length"),
                "gate_width": product.get("gate_width"),

                "inject_part": None,
                "product_type": project.get("product_type"),
                "product_total_weight": project.get("product_total_weight"),

                "product_no": project.get("product_no"),
                "product_name": project.get("product_name"),
                "product_ave_thickness": product.get("ave_thickness"),
                "product_max_thickness": product.get("max_thickness"),
                "product_max_length": product.get("flow_length"),

                "machine_id": moldflow_machine.get("id"),
                "machine_data_source": "模流",
                "machine_trademark": machine_trademark,
                "machine_serial_no": None,

                "polymer_id": moldflow_polymer.get("id"),
                "polymer_abbreviation": material_abbreviation,
                "polymer_trademark": material_trademark,
            }
            injection_volume = None
            injection_stroke = None
            VP_switch_position = None
            # 根据型腔重量/密度,计算熔胶和注射的体积, 单位是cm³
            # 是否要考虑流道重量
            cavity_weight = result_tcode.get("7112")

            if cavity_weight:
                injection_volume = float(cavity_weight) / float(polymer_info.get("melt_density"))

            # 填充控制 如果设置的是flow rate ,那么根据填充时间(是设定的injection time,还是结果的填充时间),也可以计算体积

            # 根据螺杆直径,计算注射行程
            if injection_volume and machine_info.get("injectors_info")[0].get("screw_diameter"):
                injection_stroke = injection_volume*1000/unit_conversion.getScrewArea(machine_info.get("injectors_info")[0])
            # 根据预留V/P切换位置,最大螺杆行程的20%
            if machine_info.get("injectors_info")[0].get("max_injection_stroke"):
                VP_switch_position = machine_info.get("injectors_info")[0].get("max_injection_stroke")*0.2
            # 计算熔胶终止位置,即注射起始位置
            metering_ending_position = None
            if VP_switch_position and injection_stroke:
                metering_ending_position = VP_switch_position + injection_stroke
            # 是否考虑松退，如果松退为零，那么熔胶一段位置＝熔胶终止位置
            metering_position = metering_ending_position

            # 如果没有设定绝对或者相对螺杆位置,那么注射一段,直接设置到V/P切换位置
            if not position_sec:
                process_data["inject_stage"] = 1
                position_sec = [metering_position]

            injection_pressure = []
            for i in range(process_data.get("inject_stage")):
                if result_tcode.get("1760"):
                    injection_pressure.append(round(float(result_tcode.get("1760")),2))
            inject_para_table_data = [
                # 压力按照注塑机最大注射压力
                {"label": "压力", "unit": "MPa", "sections": injection_pressure, "max": None},
                {"label": "速度", "unit": velocity_unit, "sections": velocity_sec, "max": None},
                {"label": "位置", "unit": position_unit, "sections": position_sec, "max": None}
            ]
            inject_para = {
                "injection_stage": process_data.get("inject_stage"),
                "max_injection_stage_option": 6,
                "table_data": inject_para_table_data,
                "injection_time": result_tcode.get("1610"),
                "injection_delay_time": None,
                "cooling_time": cooling_time,
            }
            holding_para_table_data = [
                {"label": "压力", "unit": "MPa" if holding_control in ["2", "1"] else "%", "sections": pressure_sec, "max": None},
                {"label": "速度", "unit": "mm/s", "sections": [None,None,None,None,None], "max": None},
                {"label": "时间", "unit": "s", "sections": time_sec, "max": None}
            ]
            holding_para = {
                "holding_stage": process_data.get("holding_stage"),
                "max_holding_stage_option": 5,
                "table_data": holding_para_table_data,
            }
            vp_switch = {
                "VP_switch_mode": "压力",
                "VP_switch_position": VP_switch_position,
                "VP_switch_time": None,
                "VP_switch_pressure": result_tcode.get("1760"),
                "VP_switch_velocity": None,
            }
            metering_para_table_data = [{
                "label": "压力",
                "unit": "kgf/cm²",
                "sections": [None,None,None,None],
            },{
                "label": "螺杆转速",
                "unit": "rpm",
                "sections": [None,None,None,None],
            },{
                "label": "背压",
                "unit": "kgf/cm²",
                "sections": [None,None,None,None],
            },{
                "label": "位置",
                "unit": "mm",
                "sections": [metering_position,None,None,None],
            }]
            table_datas = [{
                "label": "压力",
                "unit": "kgf/cm²",
                "sections": [None,None,None,None,None,None,None,None],
                "max": None,
            },{
                "label": "速度",
                "unit": "mm/s",
                "sections": [None,None,None,None,None,None,None,None],
                "max": None,
            },{
                "label": "位置",
                "unit": "mm",
                "sections": [None,None,None,None,None,None,None,None],
                "max": None,
            }]
            decompressure_paras = [{
                "label": "储前",
                "pressure": None,
                "velocity": None,
                "time": None,
                "distance": None,
            },{
                "label": "储后",
                "pressure": None,
                "velocity": None,
                "time": None,
                "distance": None,
            }]
            metering_para = {
                "metering_stage": 1,
                "max_metering_stage_option": 4,
                "table_data": metering_para_table_data,
                "decompressure_mode_before_metering": "否",
                "decompressure_mode_after_metering": "距离",
                "decompressure_paras": decompressure_paras,
                "metering_delay_time": None,
                "metering_ending_position": metering_ending_position,
            }
            if result_tcode.get("1770"):
                first_temp = round(float(result_tcode.get("1770")),0)
            first_temp = polymer_info.get("recommend_melt_temperature")
            temp_para_table_data = [{
                "label": "温度",
                "unit": "℃",
                "sections": [first_temp,first_temp-5,first_temp-10,first_temp-15,first_temp-20,None,None,None,None,None],
            }]
            temp_para = {
                "barrel_temperature_stage": 5,
                "max_barrel_temperature_stage_option": 10,
                "table_data": temp_para_table_data,
            }
            ejector_forward = {
                "ejector_forward_stage": 1,
                "max_ejector_forward_stage_option": 5,
                "table_data": table_datas,
            }
            ejector_backward = {
                "ejector_backward_stage": 1,
                "max_ejector_backward_stage_option": 5,
                "table_data": table_datas,
            }
            ejector_setting = {
                "ejector_forward": ejector_forward,
                "ejector_backward": ejector_backward,
                "ejector_mode": None,
                "ejector_start_point": None,
                "ejector_times": None,
                "ejector_stroke": None,
                "ejector_on_opening": None,
                "ejector_delay": None,
                "ejector_keep": None,
                "ejector_pause": None,
                "ejector_blow_time": None,
                "ejector_force": None,
                "set_torque": None,
            }
            mold_opening = {
                "mold_opening_stage": 1,
                "max_mold_opening_stage_option": 5,
                "table_data": table_datas,
            }
            mold_clamping = {
                "mold_clamping_stage": 1,
                "max_mold_clamping_stage_option": 5,
                "table_data": table_datas,
            }
            opening_and_clamping_mold_setting = {
                "mold_opening": mold_opening,
                "mold_clamping": mold_clamping,

                "set_mold_clamping_force": None,
                "using_robot": None,
                "using_tool": None,
                "reset_method": None,
                "set_mold_protect_time": None,
                "set_mold_protect_velocity": None,
                "set_mold_protect_pressure": None,
                "set_mold_protect_distance": None,
                "opening_position_deviation": None,

                "turnable_method": None,
                "turnable_velocity": None,
            }
            process_detail = {
                "title": "",
                "name": "",
                "inject_para": inject_para,
                "holding_para": holding_para,
                "VP_switch": vp_switch,
                "metering_para": metering_para,
                "temp_para": temp_para,
                "ejector_setting": ejector_setting,
                "opening_and_clamping_mold_setting": opening_and_clamping_mold_setting,
            }
            process_record = {
                "process_index_id": process.get("id"),
                "precondition": precondition,
                "process_detail": process_detail,
            }
            add_process_record(process_record)

            return {"mold_flow_data":params}
        else:
            raise BizException(ERROR_ERROR_FILE_TYPE)


def get_array(txt_path, language="UTF-8", tcode=None):
    language = global_language_code
    i=0
    lines = []
    for line in open(txt_path, encoding=language):
        if tcode in line:
            i += 1
        if i == 1:
            lines.append(line)
            if "End" in line:
                break
    values = []
    for i in range(2, len(lines), 2):
        if "Value" in lines[i]:
            values.append(float(lines[i][9:].strip()))
    return values


def import_machine(request):
    company_id = request.user.get("company_id")
    absolute_path = upload_mold(request)
    if type(absolute_path) is dict and absolute_path.get("error_message"):
         return {"error_message":absolute_path.get("error_message")}
    wb = load_workbook(absolute_path)
    if "Sheet1" not in wb.sheetnames:
        raise ERROR_TEMPLATE
    sheet = wb["Sheet1"]
    global error_message
    error_message = ""
    machine = {}
    if sheet and "注塑机模板" not in str(sheet["A1"].value):
        return {"machine":machine,"error_message":"请使用注塑机模板EXCEL"}
    machine["company_id"] = company_id

    machine["data_source"] = sheet["A5"].value if sheet["A5"].value else ""  # 注塑机数据来源
    machine["manufacturer"] = str(sheet["B5"].value) if sheet["B5"].value else ""  # 注塑机品牌
    machine["trademark"] = str(sheet["C5"].value) if sheet["C5"].value else ""  # 注塑机型号
    # machine["machine_type"] = str(sheet["D5"].value) if sheet["D5"].value else ""  # 注塑机类别

    machine["asset_no"] = str(sheet["D5"].value) if sheet["D5"].value else ""  # 资产编号
    machine["serial_no"] = str(sheet["E5"].value) if sheet["E5"].value else ""  # 设备编码
    machine["internal_id"] = str_decimal("F5", sheet["F5"].value, "F5")  # 注塑机ID
    machine["communication_interface"] =  str_decimal("G5", sheet["G5"].value, "G5")  # 通讯接口
    machine["agreement"] =  str(sheet["H5"].value)  # 协议
    machine["machine_type"] = str(sheet["I5"].value)
    machine["power_method"] = str(sheet["J5"].value)
    machine["propulsion_axis"] = str(sheet["K5"].value)

    machine["pressure_unit"] = str(sheet["A9"].value)
    machine["velocity_unit"] = str(sheet["B9"].value)
    machine["temperature_unit"] = str(sheet["C9"].value)
    machine["time_unit"] = str(sheet["D9"].value)
    machine["position_unit"] = str(sheet["E9"].value)
    machine["clamping_force_unit"] = str(sheet["F9"].value)
    machine["screw_rotation_unit"] = str(sheet["G9"].value)
    machine["power_unit"] = str(sheet["H9"].value)
    machine["backpressure_unit"] = str(sheet["I9"].value)
    machine["oc_pressure_unit"] = str(sheet["J9"].value)
    machine["oc_velocity_unit"] = str(sheet["K9"].value)

    # machine["platen_size_horizon"] = sheet["A30"].value  # 模板尺寸（横*竖）（H*V）（mm）
    # machine["platen_size_vertical"] = sheet["B30"].value  # 模板尺寸（横*竖）（H*V）（mm）
    machine["min_mold_size_horizon"] = str_decimal('A30', sheet["A30"].value,sheet['A29'].value)  # 最小容模尺寸（横*竖）（H*V）（mm）
    machine["min_mold_size_vertical"] = str_decimal('B30', sheet["B30"].value,sheet['B29'].value)  # 最小容模尺寸（横*竖）（H*V）（mm）
    machine["max_mold_size_horizon"] = str_decimal('C30', sheet["C30"].value,sheet['C29'].value)  # 最大容模尺寸（横*竖）（H*V）（mm）
    machine["max_mold_size_vertical"] = str_decimal('D30', sheet["D30"].value,sheet['D29'].value)  # 最大容模尺寸（横*竖）（H*V）（mm）
    machine["min_mold_thickness"] = str_decimal('E30', sheet["E30"].value,sheet['E29'].value)  # 最小容模厚度(mm)
    machine["max_mold_thickness"] = str_decimal('F30', sheet["F30"].value,sheet['F29'].value)  # 最大容模厚度(mm)
    machine["min_platen_opening"] = str_decimal('G30', sheet["G30"].value,sheet['G29'].value)  # 模板最小开距(mm)
    machine["max_platen_opening"] = str_decimal('H30', sheet["H30"].value,sheet['H29'].value)  # 模板最大开距(mm)
    machine["locate_ring_diameter"] = str_decimal('I30', sheet["I30"].value,sheet['I29'].value)  # 定位圈/法兰孔直径

    machine["pull_rod_size"] = str(sheet["A33"].value)  # 拉杆连接头尺寸
    machine["pull_rod_diameter"] = str_decimal('B33', sheet["B33"].value,sheet['B32'].value)  # 拉杆直径
    machine["pull_rod_distance_horizon"] = str_decimal('C33', sheet["C33"].value,sheet['C32'].value)  # 拉杆间距（横*竖）（H*V）(mm)
    machine["pull_rod_distance_vertical"] = str_decimal('D33', sheet["D33"].value,sheet['D32'].value)  # 拉杆间距（横*竖）（H*V）(mm)

    machine["clamping_method"] = str(sheet["A36"].value)  # 锁模方式
    # machine["max_opening_force"] = str_decimal('B36', sheet["B36"].value,sheet['B35'].value)  # 最大开模力
    machine["max_clamping_force"] = str_decimal('B36', sheet["B36"].value,sheet['B35'].value)  # 最大锁模力(Ton)
    machine["max_mold_open_stroke"] = str_decimal('C36', sheet["C36"].value,sheet['C35'].value)  # 开模行程(mm)
    # machine["max_clamping_velocity"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 最大合模速度
    # machine["max_opening_velocity"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 最大开模速度

    machine["max_ejection_force"] = str_decimal('A39', sheet["A39"].value,sheet['A38'].value)  # 顶出力
    machine["max_ejection_stroke"] = str_decimal('B39', sheet["B39"].value,sheet['B38'].value)  # 顶出行程
    machine["ejection_hole_num"] = str_decimal('C39', sheet["C39"].value,sheet['C38'].value)  # 顶出孔数量
    # machine["max_thimble_forward_speed"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 顶针最大顶进速度
    # machine["max_thimble_back_speed"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 顶针最大顶退速度

    machine["hydraulic_system_pressure"] = str_decimal('A43', sheet["A43"].value,sheet['A42'].value)  # 最大系统压力
    machine["motor_power"] = str_decimal('B43', sheet["B43"].value,sheet['B42'].value)  # 总功率电机（KW）
    machine["heater_power"] = str_decimal('C43', sheet["C43"].value,sheet['C42'].value)  # 电热功率
    machine["temp_control_zone_num"] = str_decimal('D43', sheet["D43"].value,sheet['D42'].value)  # 温度控制区数
    # machine["main_power"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 工作电压（V）
    machine["needle_core"] = str(sheet["E43"].value)  if sheet["E43"].value else ""  # 抽芯
    machine["core_pulling"] = str(sheet["F43"].value) if sheet["F43"].value else ""  # 抽芯组数
    # machine["power_method"] = str(sheet["G43"].value) if sheet["G43"].value else ""  # 动力方式

    machine["machine_weight"] = str_decimal('A47', sheet["A47"].value,sheet['A46'].value)  # 机器重量
    machine["size_length"] = str_decimal('B47', sheet["B47"].value,sheet['B46'].value)  # 机台外形尺寸（L*W*H）(mm)
    machine["size_width"] = str_decimal('C47', sheet["C47"].value,sheet['C46'].value)  # 机台外形尺寸（L*W*H）(mm)
    machine["size_height"] = str_decimal('D47', sheet["D47"].value,sheet['D46'].value)  # 机台外形尺寸（L*W*H）(mm)
    # machine["hopper_capacity"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 料斗容积
    # machine["core_pulling"] = str_decimal('A5', sheet["A5"].value,sheet['A4'].value)  # 抽芯（组）
    machine["response_time"] = str_decimal('E47', sheet["E47"].value,sheet['E46'].value)  # 响应时间
    machine["enhancement_ratio"] = str(sheet["F47"].value) if sheet["F47"].value else ""  # 增强比
    machine["manufacturing_date"] = sheet["G47"].value.date() if sheet["G47"].value else None  # 制造日期   
    machine["manufacture_date"] = sheet["H47"].value.date() if sheet["H47"].value else None  # 出厂日期
    machine["manufacture_no"] = str(sheet["I47"].value)  if sheet["I47"].value else ""  # 出厂编码
    machine["remark"] = str(sheet["J47"].value) if sheet["J47"].value else ""  # 备注

    injectors_info = {}  # 注射部件信息
    injectors_info["title"] = "部件1"
    injectors_info["name"] = "1"
    injectors_info["serial_no"] = str(sheet["B12"].value) if sheet["B12"].value else ""  # 射台编码

    injectors_info["nozzle_type"] = str(sheet["A14"].value) if sheet["A14"].value else ""  # 喷嘴类型
    injectors_info["nozzle_protrusion"] = str_decimal('B14', sheet["B14"].value,sheet['B13'].value)  # 喷嘴伸出量
    injectors_info["nozzle_hole_diameter"] = str_decimal("C14", sheet["C14"].value, sheet["C13"].value)  # 喷嘴孔直径
    injectors_info["nozzle_sphere_diameter"] = str_decimal("D14", sheet["D14"].value, sheet["D13"].value)  # 喷嘴球半径（SR）
    injectors_info["nozzle_force"] = str_decimal('E14', sheet["E14"].value,sheet['E13'].value)  # 喷嘴接触力

    injectors_info["screw_type"] = str(sheet["A17"].value) if sheet["A17"].value else ""  # 螺杆类别
    injectors_info["screw_diameter"] = str_decimal('B17', sheet["B17"].value,sheet['B16'].value)  # 螺杆直径(mm)
    # injectors_info["screw_length"] = str_decimal('C17', sheet["C17"].value,sheet['C16'].value)  # 螺杆长度(mm)
    injectors_info["screw_length_diameter_ratio"] = str_decimal('C17', sheet["C17"].value,sheet['C16'].value)  # 螺杆长径比L/D
    injectors_info["screw_compression_ratio"] = str_decimal('D17', sheet["D17"].value,sheet['D16'].value)  # 螺杆压缩比
    injectors_info["plasticizing_capacity"] = str_decimal('E17', sheet["E17"].value,sheet['E16'].value)  # 塑化能力
    # injectors_info["barrel_heating_sections"] = str_decimal('F17', sheet["F17"].value,sheet['F16'].value)  # 料筒加热段数
    injectors_info["barrel_heating_power"] = str_decimal('F17', sheet["F17"].value,sheet['F16'].value)  # 料筒加热功率
    injectors_info["max_injection_volume"] = str_decimal('G17', sheet["G17"].value,sheet['G16'].value)  # 最大注射容积
    injectors_info["max_injection_weight"] = str_decimal('H17', sheet["H17"].value,sheet['H16'].value)  # 最大注射重量
    injectors_info["max_injection_stroke"] = str_decimal('I17', sheet["I17"].value,sheet['I16'].value)  # 最大注射行程
    # injectors_info["cylinder_numer"] = str_decimal('K17', sheet["K17"].value,sheet['K16'].value)  # 油缸数
    # injectors_info["cylinder_diameter"] = str_decimal('L17', sheet["L17"].value,sheet['L16'].value)  # 油缸直径
    # injectors_info["piston_rod_diameter"] = str_decimal('M17', sheet["M17"].value,sheet['M16'].value)  # 活塞杆位于注射侧
    # injectors_info["use_small_size"] = str_decimal('N17', sheet["N17"].value,sheet['N16'].value)  # 活塞杆直径

    injectors_info["max_injection_pressure"] = str_decimal('A20', sheet["A20"].value,sheet['A19'].value)  # 最大注射压力(MPa)
    injectors_info["max_injection_velocity"] = str_decimal('B20', sheet["B20"].value,sheet['B19'].value)  # 最大注射速度(mm/s)
    injectors_info["max_holding_pressure"] = str_decimal('C20', sheet["C20"].value,sheet['C19'].value)  # 最大保压压力
    injectors_info["max_holding_velocity"] = str_decimal('D20', sheet["D20"].value,sheet['D19'].value)  # 最大保压速度(mm/s)
    injectors_info["max_metering_pressure"] = str_decimal('E20', sheet["E20"].value,sheet['E19'].value)  # 最大计量压力
    injectors_info["max_screw_rotation_speed"] = str_decimal('F20', sheet["F20"].value,sheet['F19'].value)  # 最大螺杆转速
    injectors_info["max_metering_back_pressure"] = str_decimal('G20', sheet["G20"].value,sheet['G19'].value)  # 最大计量背压
    injectors_info["max_decompression_pressure"] = str_decimal('H20', sheet["H20"].value,sheet['H19'].value)  # 最大松退压力
    injectors_info["max_decompression_velocity"] = str_decimal('I20', sheet["I20"].value,sheet['I19'].value)  # 最大松退速度
    injectors_info["max_ejector_forward_velocity"] = str_decimal('J20', sheet["J20"].value,sheet['J19'].value)  # 最大顶进速度
    injectors_info["max_ejector_backward_velocity"] = str_decimal('K20', sheet["K20"].value,sheet['K19'].value)  # 最大顶退速度
    injectors_info["max_mold_opening_velocity"] = str_decimal('L20', sheet["L20"].value,sheet['L19'].value)  # 最大开模速度
    injectors_info["max_mold_clamping_velocity"] = str_decimal('M20', sheet["M20"].value,sheet['M19'].value)  # 最大合模速度

    injectors_info["max_set_injection_pressure"] = str_decimal('A23', sheet["A23"].value,sheet['A22'].value)  # 最大可设定注射压力
    injectors_info["max_set_injection_velocity"] = str_decimal('B23', sheet["B23"].value,sheet['B22'].value)  # 最大可设定注射速度
    injectors_info["max_set_holding_pressure"] = str_decimal('C23', sheet["C23"].value,sheet['C22'].value)  # 最大可设定保压压力
    injectors_info["max_set_holding_velocity"] = str_decimal('D23', sheet["D23"].value,sheet['D22'].value)  # 最大可设定保压速度
    injectors_info["max_set_metering_pressure"] = str_decimal('E23', sheet["E23"].value,sheet['E22'].value)  # 最大可设定计量压力
    injectors_info["max_set_screw_rotation_speed"] = str_decimal('F23', sheet["F23"].value,sheet['F22'].value)  # 最大可设定螺杆转速
    injectors_info["max_set_metering_back_pressure"] = str_decimal('G23', sheet["G23"].value,sheet['G22'].value)  # 最大可设定计量背压
    injectors_info["max_set_decompression_pressure"] = str_decimal('H23', sheet["H23"].value,sheet['H22'].value)  # 最大可设定松退压力
    injectors_info["max_set_decompression_velocity"] = str_decimal('I23', sheet["I23"].value,sheet['I22'].value)  # 最大可设定松退速度
    injectors_info["max_set_ejector_forward_velocity"] = str_decimal('J23', sheet["J23"].value,sheet['J22'].value)  # 最大可设定顶进速度
    injectors_info["max_set_ejector_backward_velocity"] = str_decimal('K23', sheet["K23"].value,sheet['K22'].value)  # 最大可设定顶退速度
    injectors_info["max_set_mold_opening_velocity"] = str_decimal('L23', sheet["L23"].value,sheet['L22'].value)  # 最大可设定开模速度
    injectors_info["max_set_mold_clamping_velocity"] = str_decimal('M23', sheet["M23"].value,sheet['M22'].value)  # 最大可设定合模速度

    injectors_info["max_injection_stage"] = str_decimal('A26', sheet["A26"].value,sheet['A25'].value)  # 注射段数
    injectors_info["max_holding_stage"] = str_decimal('B26', sheet["B26"].value,sheet['B25'].value)  # 保压段数
    injectors_info["max_metering_stage"] = str_decimal('C26', sheet["C26"].value,sheet['C25'].value)  # 计量段数
    injectors_info["max_temperature_stage"] = str_decimal('D26', sheet["D26"].value,sheet['D25'].value)  # 料筒加热段数
    injectors_info["max_opening_and_clamping_stage"] = str_decimal('E26', sheet["E26"].value,sheet['E25'].value)  # 开合模设定段数
    injectors_info["max_ejector_stage"] = str_decimal('F26', sheet["F26"].value,sheet['F25'].value)  # 顶针设定段数

    machine["injectors_info"] = [injectors_info]
    return {"machine":machine,"error_message":error_message}


def str_decimal(location, value, col):
    global error_message
    if type(value) is int or type(value) is float:
        return value
    # 用正则过滤出数字,有小数点:不符合要求的String不做转化,返回None
    pattern = re.compile(r"(\d.*\d)")
    if value:
        return_value = pattern.findall(value)
        if len(return_value) > 0:
            return decimal.Decimal(return_value[0])
    col = col if col is not None else ""
    if value and type(value) != "NoneType" and value != "":
        error_message += location + " " + col  + "须为数字,不能是字符串,已自动置为空<br />"
    return None


def import_mold(request):
    company_id = request.user.get("company_id")
    absolute_path = upload_mold(request)
    if type(absolute_path) is dict and absolute_path.get("error_message"):
         return {"error_message":absolute_path.get("error_message")}
    wb = load_workbook(absolute_path)
    if "Sheet1" not in wb.sheetnames:
        raise ERROR_TEMPLATE
    sheet = wb["Sheet1"]
    if not sheet or "模具模板" not in str(sheet["A1"].value):
        return {"error_message":"请使用模具模板EXCEL"}
    global error_message
    error_message = ''
    project = Project()
    project.company_id = company_id
    project.status = 1
    if sheet["A5"] == "":
        error_message = "模号为空,请输入模号<br /><br />"
    if project_service.get_project_obj_by_mold_no(sheet["A5"].value, company_id):
        error_message = "模号已存在,请修改模号<br /><br />"
    project.mold_no = str(sheet["A5"].value) if sheet["A5"].value else ""
    project.mold_type = str(sheet["B5"].value) if sheet["B5"].value else ""
    project.mold_name = str(sheet["C5"].value) if sheet["C5"].value else ""
    project.cavity_num = str(sheet["D5"].value) if sheet["D5"].value else ""
    project.inject_cycle_require = str(sheet["E5"].value) if sheet["E5"].value else ""
    project.subrule_no = str(sheet["F5"].value) if sheet["F5"].value else ""

    project.product_category = str(sheet["A9"].value) if sheet["A9"].value else ""
    project.product_type = str(sheet["B9"].value) if sheet["B9"].value else ""
    project.product_small_type = str(sheet["C9"].value) if sheet["C9"].value else ""
    project.product_name = str(sheet["D9"].value) if sheet["D9"].value else ""
    project.product_no = str(sheet["E9"].value) if sheet["E9"].value else ""
    
    project.product_total_weight = str_decimal('F9',sheet["F9"].value, sheet["F8"].value)
    project.product_projected_area = str_decimal('G9',sheet["G9"].value, sheet["G8"].value)

    project.cavity_cooling_water_diameter = str_decimal('A18',sheet["A18"].value, sheet["A17"].value)
    project.cavity_cooling_circuit_number = str_decimal('B18',sheet["B18"].value, sheet["B17"].value)
    project.cavity_water_nozzle_specification = str_decimal('C18',sheet["C18"].value, sheet["C17"].value)
    project.core_cooling_water_diameter = str_decimal('A21', sheet["A21"].value, sheet["A20"].value)
    project.core_cooling_circuit_number = str_decimal('B21', sheet["B21"].value, sheet["B20"].value)
    project.core_water_nozzle_specification = str_decimal('C21', sheet["C21"].value, sheet["C20"].value)

    project.ejector_stroke = str_decimal('A25', sheet["A25"].value, sheet["A24"].value)
    project.ejector_rod_hole_diameter = str_decimal('B25', sheet["B25"].value, sheet["B24"].value)
    project.ejector_rod_hole_spacing = str_decimal('C25', sheet["C25"].value, sheet["C24"].value)
    project.ejector_rod_number = str_decimal('D25', sheet["D25"].value, sheet["D24"].value)
    project.ejector_force = str_decimal('E25', sheet["E25"].value, sheet["E24"].value)
    project.ejector_times = str_decimal('F25', sheet["F25"].value, sheet["F24"].value)
    project.reset_method = str(sheet["G25"].value) if sheet["G25"].value else ""
    project.ejection_method = str(sheet["H25"].value) if sheet["H25"].value else ""
    project.ejector_position_length = str_decimal('I25', sheet["I25"].value, sheet["I24"].value)
    project.ejector_position_width = str_decimal('J25', sheet["J25"].value, sheet["J24"].value)
    project.drain_distance = str_decimal('K25', sheet["K25"].value, sheet["K24"].value)

    project.mold_weight = str_decimal('A29', sheet["A29"].value, sheet["A28"].value)
    project.hanging_mold_hole_specification = str(sheet["B29"].value) if sheet["B29"].value else ""

    project.size_horizon = str_decimal('A33', sheet["A33"].value, sheet["A32"].value)
    project.size_vertical = str_decimal('B33', sheet["B33"].value, sheet["B32"].value)
    project.size_thickness = str_decimal('C33', sheet["C33"].value, sheet["C32"].value)
    project.min_clamping_force = str_decimal('D33', sheet["D33"].value, sheet["D32"].value)
    project.mold_opening_stroke = str_decimal('E33', sheet["E33"].value, sheet["E32"].value)
    project.locate_ring_diameter = str_decimal('F33', sheet["F33"].value, sheet["F32"].value)

    project.customer = str(sheet["A37"].value) if sheet["A37"].value else ""
    project.project_engineer = str(sheet["B37"].value) if sheet["B37"].value else ""
    project.design_engineer = str(sheet["C37"].value) if sheet["C37"].value else ""
    project.production_engineer = str(sheet["D37"].value) if sheet["D37"].value else ""
    project.order_date = sheet["E37"].value.date() if sheet["E37"].value else None
    product_infos = []
    loop_num = 1 if "单色模" in project.mold_type else 2 if "双色模" in project.mold_type else 3
    for index in range(loop_num):
        product = import_product(sheet, index)
        product_infos.append(product)
    assisting_equipments = []
    for col in range(10):
        if sheet.cell(33, 7+col).value == "是":
            assisting_equipments.append(sheet.cell(32, 7+col).value)
    project.assisting_equipments = "|".join(assisting_equipments)

    mold_info = project.to_dict()
    mold_info["product_infos"] = product_infos

    return {"mold":mold_info,"error_message":error_message}


def import_product(sheet, index):
    product = Product()

    product.ave_thickness = str_decimal("H"+str(9+index), sheet["H"+str(9+index)].value, sheet["H8"].value)
    product.max_thickness = str_decimal('I'+str(9+index),sheet["I"+str(9+index)].value, sheet["I8"].value)
    product.flow_length = str_decimal('J'+str(9+index),sheet["J"+str(9+index)].value, sheet["J8"].value)
    product.single_volume = str_decimal('K'+str(9+index),sheet["K"+str(9+index)].value, sheet["K8"].value)
    product.single_weight = str_decimal('L'+str(9+index),sheet["L"+str(9+index)].value, sheet["L8"].value)

    # product.locate_ring_diameter = str(sheet["A"+str(13+index)].value) if sheet["A"+str(13+index)].value else ""
    product.sprue_hole_diameter = str_decimal('A13',sheet["A"+str(13+index)].value, sheet["A12"].value)
    product.sprue_sphere_radius = str_decimal('B13',sheet["B"+str(13+index)].value, sheet["B12"].value)
    product.runner_type = str(sheet["C"+str(13+index)].value) if sheet["C"+str(13+index)].value else ""
    product.runner_length = str_decimal('D13',sheet["D"+str(13+index)].value, sheet["D12"].value)
    product.runner_weight = str_decimal('E13',sheet["E"+str(13+index)].value, sheet["E12"].value)
    product.gate_type = str(sheet["F"+str(13+index)].value) if sheet["F"+str(13+index)].value else ""
    product.gate_num = str_decimal('G13',sheet["G"+str(13+index)].value, sheet["G12"].value)
    product.gate_shape = str(sheet["H"+str(13+index)].value) if sheet["H"+str(13+index)].value else ""
    product.gate_area = str_decimal('I13',sheet["I"+str(13+index)].value, sheet["I12"].value)
    product.gate_radius = str_decimal('J13',sheet["J"+str(13+index)].value, sheet["J12"].value)
    product.gate_length = str_decimal('K13',sheet["K"+str(13+index)].value, sheet["K12"].value)
    product.gate_width = str_decimal('L13',sheet["L"+str(13+index)].value, sheet["L12"].value)
    product.valve_num = str_decimal('M13',sheet["M"+str(13+index)].value, sheet["M12"].value)
    product.hot_runner_num = str_decimal('N13',sheet["N"+str(13+index)].value, sheet["N12"].value)
    return product.to_dict()

def import_polymer(request):
    company_id = request.user.get("company_id")
    absolute_path = upload_mold(request)
    if type(absolute_path) is dict and absolute_path.get("error_message"):
         return {"error_message":absolute_path.get("error_message")}
    wb = load_workbook(absolute_path)
    if "Sheet1" not in wb.sheetnames:
        raise ERROR_TEMPLATE
    sheet = wb["Sheet1"]
    global error_message
    error_message = ""
    polymer = {}
    if "材料模板" not in str(sheet["A1"].value):
        return {"polymer":polymer,"error_message":"请使用材料模板EXCEL"}
    polymer["company_id"] = company_id
    polymer["abbreviation"] = str(sheet["A5"].value) if sheet["A5"].value else ""
    polymer["trademark"] = str(sheet["B5"].value) if sheet["B5"].value else ""
    polymer["manufacturer"] = str(sheet["C5"].value) if sheet["C5"].value else ""
    polymer["category"] = str(sheet["D5"].value) if sheet["D5"].value else ""
    polymer["data_source"] = str(sheet["E5"].value) if sheet["E5"].value else ""
    polymer["data_status"] = str(sheet["F5"].value) if sheet["F5"].value else ""
    polymer["internal_id"] = str(sheet["G5"].value) if sheet["G5"].value else ""
    polymer["level_code"] = str(sheet["H5"].value) if sheet["H5"].value else ""
    polymer["vendor_code"] = str(sheet["I5"].value) if sheet["I5"].value else ""

    polymer["max_melt_temperature"] = str_decimal("A9", sheet["A9"].value,sheet["A8"].value)
    polymer["min_melt_temperature"] = str_decimal("B9", sheet["B9"].value,sheet["B8"].value)
    polymer["recommend_melt_temperature"] = str_decimal("C9", sheet["C9"].value,sheet["C8"].value)
    polymer["max_mold_temperature"] = str_decimal("D9", sheet["D9"].value,sheet["D8"].value)
    polymer["min_mold_temperature"] = str_decimal("E9", sheet["E9"].value,sheet["E8"].value)
    polymer["recommend_mold_temperature"] = str_decimal("F9", sheet["F9"].value,sheet["F8"].value)
    polymer["max_shear_linear_speed"] = str_decimal("G9", sheet["G9"].value,sheet["G8"].value)
    polymer["min_shear_linear_speed"] = str_decimal("H9", sheet["H9"].value,sheet["H8"].value)
    polymer["recommend_shear_linear_speed"] = str_decimal("I9", sheet["I9"].value,sheet["I8"].value)
    polymer["degradation_temperature"] = str_decimal("J9", sheet["J9"].value,sheet["J8"].value)
    polymer["ejection_temperature"] = str_decimal("K9", sheet["K9"].value,sheet["K8"].value)
    polymer["recommend_injection_rate"] = str_decimal("L9", sheet["L9"].value,sheet["L8"].value)
    polymer["max_sheer_rate"] = str_decimal("A11", sheet["A11"].value,sheet["A10"].value)
    polymer["max_sheer_stress"] = str_decimal("B11", sheet["B11"].value,sheet["B10"].value)
    polymer["recommend_back_pressure"] = str_decimal("C11", sheet["C11"].value,sheet["C10"].value)
    polymer["barrel_residence_time"] = str_decimal("D11", sheet["D11"].value,sheet["D10"].value)
    polymer["dry_method"] = str(sheet["A14"].value) if sheet["A14"].value else ""
    polymer["dry_temperature"] = str(sheet["B14"].value) if sheet["B14"].value else ""
    polymer["dry_time"] = str(sheet["C14"].value) if sheet["C14"].value else ""

    polymer["viscosity_model"] = "cross_WLF"
    polymer["cross_WLF_n"] = str(sheet["A25"].value) if sheet["A25"].value else ""
    polymer["cross_WLF_Tau"] = str(sheet["B25"].value) if sheet["B25"].value else ""
    polymer["cross_WLF_D1"] = str(sheet["C25"].value) if sheet["C25"].value else ""
    polymer["cross_WLF_D2"] = str(sheet["D25"].value) if sheet["D25"].value else ""
    polymer["cross_WLF_D3"] = str(sheet["E25"].value) if sheet["E25"].value else ""
    polymer["cross_WLF_A1"] = str(sheet["F25"].value) if sheet["F25"].value else ""
    polymer["cross_WLF_A2"] = str(sheet["G25"].value) if sheet["G25"].value else ""
    polymer["c1"] = str(sheet["A28"].value) if sheet["A28"].value else ""
    polymer["c2"] = str(sheet["B28"].value) if sheet["B28"].value else ""
    polymer["switch_temp"] = str(sheet["C28"].value) if sheet["C28"].value else ""
    polymer["viscosity_index"] = str(sheet["D28"].value) if sheet["D28"].value else ""
    polymer["MFR_temp"] = str(sheet["E28"].value) if sheet["E28"].value else ""
    polymer["MFR_load"] = str(sheet["F28"].value) if sheet["F28"].value else ""
    polymer["MFR_measure"] = str(sheet["G28"].value) if sheet["G28"].value else ""

    polymer["melt_density"] = str_decimal("A18", sheet["A18"].value,sheet["A17"].value)
    polymer["solid_density"] = str_decimal("B18", sheet["B18"].value,sheet["B17"].value)
    polymer["Tait_pvT_b5"] = str(sheet["C18"].value) if sheet["C18"].value else ""
    polymer["Tait_pvT_b6"] = str(sheet["D18"].value) if sheet["D18"].value else ""
    polymer["Tait_pvT_b1m"] = str(sheet["E18"].value) if sheet["E18"].value else ""
    polymer["Tait_pvT_b2m"] = str(sheet["F18"].value) if sheet["F18"].value else ""
    polymer["Tait_pvT_b3m"] = str(sheet["G18"].value) if sheet["G18"].value else ""
    polymer["Tait_pvT_b4m"] = str(sheet["H18"].value) if sheet["H18"].value else ""
    polymer["Tait_pvT_b1s"] = str(sheet["I18"].value) if sheet["I18"].value else ""
    polymer["Tait_pvT_b2s"] = str(sheet["J18"].value) if sheet["J18"].value else ""
    polymer["Tait_pvT_b3s"] = str(sheet["K18"].value) if sheet["K18"].value else ""
    polymer["Tait_pvT_b4s"] = str(sheet["L18"].value) if sheet["L18"].value else ""
    polymer["Tait_pvT_b7"] = str(sheet["A20"].value) if sheet["A20"].value else ""
    polymer["Tait_pvT_b8"] = str(sheet["B20"].value) if sheet["B20"].value else ""
    polymer["Tait_pvT_b9"] = str(sheet["C20"].value) if sheet["C20"].value else ""

    polymer["E1"] = str_decimal("A33", sheet["A33"].value,sheet["A32"].value)
    polymer["E2"] = str_decimal("B33", sheet["B33"].value,sheet["B32"].value)
    polymer["v12"] = str_decimal("C33", sheet["C33"].value,sheet["C32"].value)
    polymer["v23"] = str_decimal("D33", sheet["D33"].value,sheet["D32"].value)
    polymer["G12"] = str_decimal("E33", sheet["E33"].value,sheet["E32"].value)
    polymer["Alpha1"] = str_decimal("A36", sheet["A36"].value,sheet["A35"].value)
    polymer["Alpha2"] = str_decimal("B36", sheet["B36"].value,sheet["B35"].value)

    polymer["average_horizontal_shrinkage"] = str_decimal("A41", sheet["A41"].value,sheet["A40"].value)
    polymer["average_vertical_shrinkage"] = str_decimal("B41", sheet["B41"].value,sheet["B40"].value)
    polymer["min_horizontal_shrinkage"] = str_decimal("A44", sheet["A44"].value,sheet["A43"].value)
    polymer["max_horizontal_shrinkage"] = str_decimal("B44", sheet["B44"].value,sheet["B43"].value)
    polymer["min_vertical_shrinkage"] = str_decimal("C44", sheet["C44"].value,sheet["C43"].value)
    polymer["max_vertical_shrinkage"] = str_decimal("D44", sheet["D44"].value,sheet["D43"].value)

    polymer["filler"] = str(sheet["A49"].value) if sheet["A49"].value else ""
    polymer["filler_type"] = str(sheet["B49"].value) if sheet["B49"].value else ""
    polymer["filler_shape"] = str(sheet["C49"].value) if sheet["C49"].value else ""
    polymer["filler_percentage"] = str_decimal("D49", sheet["D49"].value,sheet["D48"].value)
    polymer["filler_density"] = str_decimal("E49", sheet["E49"].value,sheet["E48"].value)
    polymer["filler_specific_heat"] = str_decimal("F49", sheet["F49"].value,sheet["F48"].value)
    polymer["filler_specific_thermal_conductivity"] = str_decimal("G49", sheet["G49"].value,sheet["G48"].value)
    polymer["filler_E1"] = str_decimal("A52", sheet["A52"].value,sheet["A51"].value)
    polymer["filler_E2"] = str_decimal("B52", sheet["B52"].value,sheet["B51"].value)
    polymer["filler_v12"] = str_decimal("C52", sheet["C52"].value,sheet["C51"].value)
    polymer["filler_v23"] = str_decimal("D52", sheet["D52"].value,sheet["D51"].value)
    polymer["filler_G12"] = str_decimal("E52", sheet["E52"].value,sheet["E51"].value)
    polymer["filler_Alpha1"] = str_decimal("A55", sheet["A55"].value,sheet["A54"].value)
    polymer["filler_Alpha2"] = str_decimal("B55", sheet["B55"].value,sheet["B54"].value)
    polymer["filler_horizontal_tensile_strength"] = str_decimal("A59", sheet["A59"].value,sheet["A58"].value)
    polymer["filler_vertical_tensile_strength"] = str_decimal("B59", sheet["B59"].value,sheet["B58"].value)
    polymer["filler_aspect_ratio"] = str_decimal("C59", sheet["C59"].value,sheet["C58"].value)

    return {"polymer":polymer,"error_message":error_message}


def import_process_record(request):
    company_id = request.user.get("company_id")
    absolute_path = upload_mold(request)
    wb = load_workbook(absolute_path)
    sheet = wb["Sheet1"]
    global error_message
    error_message = ""
    process_record = {}
    process_index_id = 0
    precondition = {}
    process_detail= {}
    if "工艺导出模板" not in str(sheet["A1"].value):
        return {"process_record":process_record,"error_message":"请使用工艺导出模板EXCEL"}
    precondition["mold_no"] = str(sheet["A10"].value)
    precondition["cavity_num"] = str(sheet["B10"].value)
    precondition["runner_length"] = str_decimal("C10", sheet["C10"].value,sheet["C9"].value)
    precondition["runner_weight"] = str_decimal("D10", sheet["D10"].value,sheet["C9"].value)
    precondition["gate_type"] = str(sheet["E10"].value)
    precondition["gate_num"] = str_decimal("F10", sheet["F10"].value,sheet["F9"].value)
    precondition["gate_shape"] = str(sheet["G10"].value)

    precondition["product_no"] = str(sheet["A14"].value)
    precondition["product_type"] = str(sheet["B14"].value)
    precondition["product_name"] = str(sheet["C14"].value)
    precondition["product_total_weight"] = str_decimal("D14", sheet["D14"].value,sheet["D13"].value)
    precondition["product_ave_thickness"] = str_decimal("E14", sheet["E14"].value,sheet["E13"].value)
    precondition["product_max_thickness"] = str_decimal("F14", sheet["F14"].value,sheet["F13"].value)
    precondition["product_max_length"] = str_decimal("G14", sheet["G14"].value,sheet["G13"].value)

    precondition["machine_data_source"] = str(sheet["A6"].value)
    precondition["machine_trademark"] = str(sheet["B6"].value)
    precondition["machine_serial_no"] = str(sheet["C6"].value)

    precondition["polymer_abbreviation"] = str(sheet["D6"].value)
    precondition["polymer_trademark"] = str(sheet["E6"].value)

    total, machine = machine_service.get_list_of_machine(trademark=precondition["machine_trademark"])
    if total == 0:
        error_message += "注塑机型号不匹配,请重新选择注塑机<br>"
    else:
        precondition["machine_id"] = machine[0].get("id")
    total, polymer = polymer_service.get_list_of_polymer(trademark=precondition["polymer_trademark"])
    if total == 0:
        error_message += "塑料牌号不匹配,请重新选择胶料<br>"
    else:
        precondition["polymer_id"] = polymer[0].get("id")
    total, mold = project_service.get_list_of_project(mold_no=precondition["mold_no"])
    if total == 0:
        error_message += "模号不匹配,请重新选择模具编号<br>"
    else:
        precondition["mold_id"] = mold[0].get("id")

    inject_para_table_data = [{
        "label":"压力",
        "unit":"bar",
        "sections":[
            str_decimal("B20", sheet["B20"].value,sheet["B18"].value),
            str_decimal("C20", sheet["C20"].value,sheet["C18"].value),
            str_decimal("D20", sheet["D20"].value,sheet["D18"].value),
            str_decimal("E20", sheet["E20"].value,sheet["E18"].value),
            str_decimal("F20", sheet["F20"].value,sheet["F18"].value),
            str_decimal("G20", sheet["G20"].value,sheet["G18"].value)

        ]
    },{
        "label":"速度",
        "unit":"%",
        "sections":[
            str_decimal("B21", sheet["B21"].value,sheet["B18"].value),
            str_decimal("C21", sheet["C21"].value,sheet["C18"].value),
            str_decimal("D21", sheet["D21"].value,sheet["D18"].value),
            str_decimal("E21", sheet["E21"].value,sheet["E18"].value),
            str_decimal("F21", sheet["F21"].value,sheet["F18"].value),
            str_decimal("F21", sheet["F21"].value,sheet["G18"].value)

        ]
    },{
        "label":"位置",
        "unit":"mm",
        "sections":[
            str_decimal("B22", sheet["B22"].value,sheet["B18"].value),
            str_decimal("C22", sheet["C22"].value,sheet["C18"].value),
            str_decimal("D22", sheet["D22"].value,sheet["D18"].value),
            str_decimal("E22", sheet["E22"].value,sheet["E18"].value),
            str_decimal("F22", sheet["F22"].value,sheet["F18"].value),
            str_decimal("G22", sheet["G22"].value,sheet["G18"].value)

        ]
    }]
    inject_para = {}
    inject_para["injection_stage"] = 6
    inject_para["max_injection_stage_option"] = 6
    inject_para["table_data"] = inject_para_table_data
    inject_para["injection_time"] = str_decimal("A25", sheet["A25"].value,sheet["A24"].value)
    inject_para["injection_delay_time"] = str_decimal("B25", sheet["B25"].value,sheet["A24"].value)
    inject_para["cooling_time"] = str_decimal("C25", sheet["C25"].value,sheet["C24"].value)

    holding_para_table_data = [{
        "label":"压力",
        "unit":"bar",
        "sections":[
            str_decimal("B33", sheet["B33"].value,sheet["B31"].value),
            str_decimal("C33", sheet["C33"].value,sheet["C31"].value),
            str_decimal("D33", sheet["D33"].value,sheet["D31"].value),
            str_decimal("E33", sheet["E33"].value,sheet["E31"].value),
            str_decimal("F33", sheet["F33"].value,sheet["F31"].value)
        ]
    },{
        "label":"速度",
        "unit":"%",
        "sections":[
            str_decimal("B34", sheet["B34"].value,sheet["B31"].value),
            str_decimal("C34", sheet["C34"].value,sheet["C31"].value),
            str_decimal("D34", sheet["D34"].value,sheet["D31"].value),
            str_decimal("E34", sheet["E34"].value,sheet["E31"].value),
            str_decimal("F34", sheet["F34"].value,sheet["F31"].value)
        ]
    },{
        "label":"位置",
        "unit":"mm",
        "sections":[
            str_decimal("B35", sheet["B35"].value,sheet["B31"].value),
            str_decimal("C35", sheet["C35"].value,sheet["C31"].value),
            str_decimal("D35", sheet["D35"].value,sheet["D31"].value),
            str_decimal("E35", sheet["E35"].value,sheet["E31"].value),
            str_decimal("F35", sheet["F35"].value,sheet["F31"].value)
        ]
    }]
    holding_para = {}
    holding_para["holding_stage"] = 5
    holding_para["max_holding_stage_option"] = 5
    holding_para["table_data"] = holding_para_table_data

    VP_switch = {}
    VP_switch["VP_switch_mode"] = str_decimal("A29", sheet["A29"].value,sheet["A28"].value)
    VP_switch["VP_switch_position"] = str_decimal("B29", sheet["B29"].value,sheet["B28"].value)
    VP_switch["VP_switch_time"] = str_decimal("C29", sheet["C29"].value,sheet["C28"].value)
    VP_switch["VP_switch_pressure"] = str_decimal("D29", sheet["D29"].value,sheet["D28"].value)
    VP_switch["VP_switch_velocity"] = str_decimal("E29", sheet["E29"].value,sheet["E28"].value)

    metering_para_table_data = [{
        "label":"压力",
        "unit":"bar",
        "sections":[
            str_decimal("B39", sheet["B39"].value,sheet["B37"].value),
            str_decimal("C39", sheet["C39"].value,sheet["C37"].value),
            str_decimal("D39", sheet["D39"].value,sheet["D37"].value),
            str_decimal("E39", sheet["E39"].value,sheet["E37"].value),
        ]
    },{
        "label":"螺杆转速",
        "unit":"%",
        "sections":[
            str_decimal("B40", sheet["B40"].value,sheet["B37"].value),
            str_decimal("C40", sheet["C40"].value,sheet["C37"].value),
            str_decimal("D40", sheet["D40"].value,sheet["D37"].value),
            str_decimal("E40", sheet["E40"].value,sheet["E37"].value),
        ]
    },{
        "label":"背压",
        "unit":"bar",
        "sections":[
            str_decimal("B41", sheet["B41"].value,sheet["B37"].value),
            str_decimal("C41", sheet["C41"].value,sheet["C37"].value),
            str_decimal("D41", sheet["D41"].value,sheet["D37"].value),
            str_decimal("E41", sheet["E41"].value,sheet["E37"].value),
        ]
    },{
        "label":"位置",
        "unit":"mm",
        "sections":[
            str_decimal("B42", sheet["B42"].value,sheet["B37"].value),
            str_decimal("C42", sheet["C42"].value,sheet["C37"].value),
            str_decimal("D42", sheet["D42"].value,sheet["D37"].value),
            str_decimal("E42", sheet["E42"].value,sheet["E37"].value),
        ]
    }
	]
    metering_para_decompressure_paras_item = [
        {
            "label":"储前", 
            "pressure":str_decimal("B48", sheet["B48"].value,sheet["B47"].value),
            "velocity":str_decimal("C48", sheet["C48"].value,sheet["C47"].value),
            "time":str_decimal("D48", sheet["D48"].value,sheet["D47"].value),
            "distance":str_decimal("E48", sheet["E48"].value,sheet["E47"].value),
        },
        {
            "label":"储后", 
            "pressure":str_decimal("B49", sheet["B49"].value,sheet["B47"].value),
            "velocity":str_decimal("C49", sheet["C49"].value,sheet["C47"].value),
            "time":str_decimal("D49", sheet["D49"].value,sheet["D47"].value),
            "distance":str_decimal("E49", sheet["E49"].value,sheet["E47"].value),
        }
    ]
    metering_para = {}
    metering_para["metering_stage"] = 4
    metering_para["max_metering_stage_option"] = 4
    metering_para["table_data"] = metering_para_table_data
    metering_para["decompressure_mode_before_metering"] = str(sheet["A45"].value) if sheet["A45"].value else ""
    metering_para["decompressure_mode_after_metering"] = str(sheet["B45"].value) if sheet["B45"].value else ""
    metering_para["decompressure_paras"] = metering_para_decompressure_paras_item
    # setting_items_metering_para["metering_delay_time"] = 
    # setting_items_metering_para["metering_ending_position"] = 

    temp_para = {
        "barrel_temperature_stage": 10,
        "max_barrel_temperature_stage_option": 10,
        "table_data": [{
            "label": "温度",
            "unit": "℃",
            "sections":[str_decimal("B53", sheet["B53"].value,sheet["B52"].value),
                str_decimal("C53", sheet["C53"].value,sheet["C52"].value),
                str_decimal("D53", sheet["D53"].value,sheet["D52"].value),
                str_decimal("E53", sheet["E53"].value,sheet["E52"].value),
                str_decimal("F53", sheet["F53"].value,sheet["F52"].value),
                str_decimal("G53", sheet["G53"].value,sheet["G52"].value),
                str_decimal("H53", sheet["H53"].value,sheet["H52"].value),
                str_decimal("I53", sheet["I53"].value,sheet["I52"].value),
                str_decimal("J53", sheet["J53"].value,sheet["J52"].value),
                str_decimal("K53", sheet["K53"].value,sheet["K52"].value),
            ]
            },
        ]
    }
    process_detail["title"] = "射台 #1"
    process_detail["name"] = "1"
    process_detail["inject_para"] = inject_para
    process_detail["holding_para"] = holding_para
    process_detail["VP_switch"] = VP_switch
    process_detail["metering_para"] = metering_para
    process_detail["temp_para"] = temp_para

    process_record["precondition"] = precondition
    process_record["process_detail"] = process_detail
    return {"process_record": process_record, "error_message":error_message}