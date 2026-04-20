"""
导出服务
"""
from array import array
import os
import time
import logging

from django.conf import settings
from openpyxl import load_workbook

from hsmolding.services import (
    project_service,
    machine_service,
    polymer_service,
)


from PIL import JpegImagePlugin
from mdprocess.services import process_record_service, rule_service
from hsmolding.services.export_report_service import (
    apply_mapping, 
    apply_mapping_multi,
    machine_mapping, 
    injectors_info_mapping, 
    polymer_mapping,
    mold_mapping,
    product_mapping
)

JpegImagePlugin._getmp = lambda x: None
_LOGGER = logging.getLogger(__name__)

reservation_info = None
machine_info = None
polymer_info  = None

"""
根据单一职责原则（Simple-Responsibility Principle）,把导出功能放在export_service里
"""

# 查看路径是否存在
def check_path_existed(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


# 导出工程列表(模具列表)
def export_project_table(project_id_list: array):
    total, projects = project_service.get_list_of_project(project_id_list=project_id_list)
    # 读取模版文件
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_mold_list.xlsx")
    if not wb:
        return { "url": "" }
    sheet = wb["Sheet1"]
    for index, project in enumerate(projects):
        sheet["A"+str(index+2)] = str(index + 1)
        sheet["B"+str(index+2)] = project["mold_no"]
        sheet["C"+str(index+2)] = project["customer"]
        sheet["D"+str(index+2)] = project["mold_name"]
        sheet["E"+str(index+2)] = project["mold_type"]
        sheet["F"+str(index+2)] = project["product_name"]
        sheet["G"+str(index+2)] = project["project_engineer"]
        sheet["H"+str(index+2)] = project["production_engineer"]
        sheet["I"+str(index+2)] = project["order_date"]

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"temp/"
    file_name = str("project_list_") + date + ".xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)
    
    return { "url": file_relative_path }


# 导出机器列表
def export_machine_table(params):
    total, machines = machine_service.list_machine_view({"machine_id_list":params.get("machine_id_list")})
    # 读取文件
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_machine_list.xlsx")
    if not wb:
        return { "url": "" }
    sheet = wb["Sheet1"]
    row = 2
    if total > 0:
        for index in range(total):
            machine_part = machines[index]
            col = 1
            for key, value in machine_part.items():
                if key not in ["asset_no", "serial_no"]:
                    sheet.cell(row, col).value = value
                    col = col + 1
            row = row + 1

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"temp/"
    file_name = str("machine_list_") + date + ".xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 导出胶料列表
def export_polymer_table(params):
    polymers = polymer_service.list_polymer_view(params.get("polymer_id_list"))
    # 读取文件
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_polymer_list.xlsx")
    if not wb:
        return { "url": "" }
    sheet = wb["Sheet1"]
    row = 2
    for polymer in polymers:
        col = 1
        for value in polymer:
            sheet.cell(row, col).value = value
            col = col + 1
        row = row + 1

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"temp/"
    file_name = str("polymer_list_") + date + ".xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 导出模具信息（通过id）
def export_mold_by_id(id: int):
    project = project_service.get_mold_dict_by_id(id)
    return export_mold(project)
    

# 导出模具信息
def export_mold(mold_info: dict):

    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_mold.xlsx")
    if not wb:
        return { "url": "" }

    sheet = wb["Sheet1"]

    product_infos = mold_info.get("product_infos")
    apply_mapping(sheet, mold_mapping, mold_info)
    if product_infos and len(product_infos) >=1:
        for i in range(len(product_infos)):
            product_info = product_infos[i]
            apply_mapping_multi(sheet, product_mapping, product_info, row=9 + i, row2=13 + i) 
   
    if mold_info.get("assisting_equipments"):
        equipments = mold_info.get("assisting_equipments").split("|")
        for col in range(10):
            if sheet.cell(32, 7+col).value in equipments:
                sheet.cell(33, 7+col).value = "是"

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"gsid_{mold_info.get('company_id')}/temp/"
    file_name = f"mold_info_{mold_info.get('mold_no')}_{date}.xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 导出模具多色
def export_mold_multi(sheet, product_infos):
    for i in range(len(product_infos)):
        product_info = product_infos[i]
        sheet["D"+str(9+i)] = product_info.get("ave_thickness") 
        sheet["E"+str(9+i)] = product_info.get("max_thickness")
        sheet["F"+str(9+i)] = product_info.get("flow_length") 
        sheet["G"+str(9+i)] = product_info.get("single_volume") 
        sheet["H"+str(9+i)] = product_info.get("single_weight") 

        sheet["A"+str(13+i)] = product_info.get("locate_ring_diameter") 
        sheet["B"+str(13+i)] = product_info.get("sprue_hole_diameter") 
        sheet["C"+str(13+i)] = product_info.get("sprue_sphere_radius") 
        sheet["D"+str(13+i)] = product_info.get("runner_type") 
        sheet["E"+str(13+i)] = product_info.get("runner_length") 
        sheet["F"+str(13+i)] = product_info.get("runner_weight") 
        sheet["G"+str(13+i)] = product_info.get("gate_type") 
        sheet["H"+str(13+i)] = product_info.get("gate_num") 
        sheet["I"+str(13+i)] = product_info.get("gate_shape") 
        sheet["J"+str(13+i)] = product_info.get("gate_area") 
        sheet["K"+str(13+i)] = product_info.get("gate_radius") 
        sheet["L"+str(13+i)] = product_info.get("gate_length") 
        sheet["M"+str(13+i)] = product_info.get("gate_width")
        sheet["N"+str(13+i)] = product_info.get("flow_length")
        sheet["O"+str(13+i)] = product_info.get("valve_num")
        sheet["P"+str(13+i)] = product_info.get("hot_runner_num")


# 根据机器id导出excel文件
def export_machine_by_id(id):
    total, machine_list = machine_service.list_machine_injector_view({"machine_id_list":[id]})
    if total > 0:
        return export_machine(machine_list[0])


# 根据机器信息导出excel文件
def export_machine(machine: dict):
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_machine.xlsx")

    sheet = wb["Sheet1"]
    injectors_info: dict = machine.get("injectors_info")[0]
    apply_mapping(sheet, machine_mapping, machine)
    apply_mapping(sheet, injectors_info_mapping, injectors_info)

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"gsid_{machine.get('company_id')}/temp/"
    file_name = f"mac_info_{machine.get('trademark')}_{date}.xlsx"
    file_name = file_name.replace("/", "_").replace("+", "_")
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 根据id导出胶料
def export_polymer_by_id(id: int):
    polymer: dict = polymer_service.get_polymer(id)
    return export_polymer(polymer)


# 根据胶料信息导出excel文件
def export_polymer(polymer: dict):
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_polymer.xlsx")
    
    sheet = wb["Sheet1"]
    apply_mapping(sheet, polymer_mapping, polymer)
       
    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"gsid_{polymer.get('company_id')}/temp/"
    file_name = f"poly_info_{polymer.get('trademark')}_{date}.xlsx"
    file_name = file_name.replace("/", "_").replace("+", "_")
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 根据id导出工艺
def export_process_by_id(id=None, company_id=None):
    process_record = process_record_service.get_process_record(id)
    return export_process(process_record, company_id)


# 根据工艺信息导出excel
def export_process(process_record: dict, company_id: int):
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_process.xlsx")
    if process_record and process_record["precondition"]:
        precondition: dict = process_record.get("precondition")
        sheet = wb["Sheet1"]
        gate_area = None
        if precondition["gate_shape"]=="圆形":
            gate_area = precondition["gate_radius"]
        elif precondition["gate_shape"]=="矩形":
            gate_area = f'{precondition["gate_length"]} * {precondition["gate_width"]}'
        sheet["A10"] = precondition["mold_no"]
        sheet["B10"] = precondition["cavity_num"]
        sheet["C10"] = precondition["runner_length"]
        sheet["D10"] = precondition["runner_weight"]
        sheet["E10"] = precondition["gate_type"]
        sheet["F10"] = precondition["gate_num"]
        sheet["G10"] = precondition["gate_shape"]
        sheet["H10"] = gate_area
        sheet["A14"] = precondition["product_no"]
        sheet["B14"] = precondition["product_type"]
        sheet["C14"] = precondition["product_name"]
        sheet["D14"] = precondition["product_total_weight"]
        sheet["E14"] = precondition["product_ave_thickness"]
        sheet["F14"] = precondition["product_max_thickness"]
        sheet["G14"] = precondition["product_max_length"]
        sheet["A6"] = precondition["machine_data_source"]
        sheet["B6"] = precondition["machine_trademark"]
        sheet["C6"] = precondition["machine_serial_no"]
        sheet["D6"] = precondition["polymer_abbreviation"]
        sheet["E6"] = precondition["polymer_trademark"]

        if precondition.get("machine_id"):
            machine_dict: dict = machine_service.get_machine(precondition.get("machine_id"))
            sheet["H20"] = machine_dict.get("pressure_unit")
            sheet["H21"] = machine_dict.get("velocity_unit")
            sheet["H22"] = machine_dict.get("position_unit")
            sheet["G33"] = machine_dict.get("pressure_unit")
            sheet["G34"] = machine_dict.get("velocity_unit")
            sheet["G35"] = machine_dict.get("time_unit")
            sheet["F39"] = machine_dict.get("pressure_unit")
            sheet["F40"] = machine_dict.get("screw_rotation_unit")
            sheet["F41"] = machine_dict.get("pressure_unit")
            sheet["F42"] = machine_dict.get("position_unit")


    if process_record and process_record.get("process_detail"):
        process_detail = process_record.get("process_detail")  
        inject_para = process_detail.get("inject_para")
        sheet["A19"] = inject_para.get("injection_stage")
        if inject_para and inject_para.get("table_data") and inject_para.get("table_data")[0].get("sections"):
            for row in range(len(inject_para.get("table_data"))):
                for num in range(len(inject_para.get("table_data")[row].get("sections"))):
                    sheet.cell(20+row, 2+num).value = inject_para.get("table_data")[row].get("sections")[num]
        sheet["A25"] = inject_para["injection_time"]
        sheet["B25"] = inject_para["injection_delay_time"]
        sheet["C25"] = inject_para["cooling_time"]

        holding_para = process_detail.get("holding_para")
        sheet["A32"] = holding_para.get("holding_stage")
        if holding_para and holding_para.get("table_data") and holding_para.get("table_data")[0].get("sections"):
            for row in range(len(holding_para.get("table_data"))):
                for num in range(len(holding_para.get("table_data")[row].get("sections"))):
                    sheet.cell(33+row, 2+num).value = holding_para.get("table_data")[row].get("sections")[num]

        VP_switch = process_detail.get("VP_switch")
        if VP_switch:
            sheet["A29"] = VP_switch["VP_switch_mode"]
            sheet["B29"] = VP_switch["VP_switch_time"]
            sheet["C29"] = VP_switch["VP_switch_position"]
            sheet["D29"] = VP_switch["VP_switch_pressure"]
            sheet["E29"] = VP_switch["VP_switch_velocity"]

        metering_para = process_detail.get("metering_para")
        sheet["A38"] = metering_para.get("metering_stage")
        if metering_para and metering_para.get("table_data") and metering_para.get("table_data")[0].get("sections"):
            for row in range(len(metering_para.get("table_data"))):
                for num in range(len(metering_para.get("table_data")[row].get("sections"))):
                    sheet.cell(39+row, 2+num).value = metering_para.get("table_data")[row].get("sections")[num]

        metering_para_decompressure_paras_item = metering_para.get("decompressure_paras")
        if metering_para_decompressure_paras_item and len(metering_para_decompressure_paras_item) > 0:
            sheet["B48"] = metering_para_decompressure_paras_item[0].get("pressure")
            sheet["C48"] = metering_para_decompressure_paras_item[0].get("velocity")
            sheet["D48"] = metering_para_decompressure_paras_item[0].get("distance")
            sheet["E48"] = metering_para_decompressure_paras_item[0].get("time")

        if metering_para_decompressure_paras_item and len(metering_para_decompressure_paras_item) > 1:
            sheet["B49"] = metering_para_decompressure_paras_item[1].get("pressure")
            sheet["C49"] = metering_para_decompressure_paras_item[1].get("velocity")
            sheet["D49"] = metering_para_decompressure_paras_item[1].get("distance")
            sheet["E49"] = metering_para_decompressure_paras_item[1].get("time")
        sheet["A45"] = metering_para["decompressure_mode_before_metering"]
        sheet["B45"] = metering_para["decompressure_mode_after_metering"]

        temp_para = process_detail.get("temp_para")
        sheet["A52"] = temp_para.get("barrel_temperature_stage")
        if temp_para and temp_para.get("table_data") and temp_para.get("table_data")[0].get("sections"):
            for num in range(len(temp_para.get("table_data")[0].get("sections"))):
                sheet.cell(53, 2+num).value = temp_para.get("table_data")[0].get("sections")[num]

    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"gsid_{company_id}/temp/"
    file_name = f"process_info_{process_record.get('precondition').get('mold_no')}_{date}.xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


# 导出规则
def export_rule(param_dict, company_id):
    subrule_no = param_dict.get("subrule_no")
    flag = param_dict.get("flag")
    _, rules = rule_service.get_list_of_rule_method(subrule_no=subrule_no)
    if rules:
        rule = rules[0]
        defect_names = set()
        folder_relative_path = f"gsid_{company_id}/temp/rule.txt"
        check_path_existed(f"{settings.FILE_STORAGE_PATH}gsid_{company_id}/temp/")
        with open(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}", 'w', encoding='utf-8') as file:
            if flag == "has_explanation":
                file.write('# 规则库编号： ' + subrule_no + '\n')
                if rule.get("product_small_type"):
                    file.write('# 制品类别：' + rule.get("product_small_type") + '\n')
                if rule.get("polymer_abbreviation"):
                    file.write('# 塑料简称：' + rule.get("polymer_abbreviation") + '\n')
                file.write('\n')
            for rule in rules:
                if flag == "has_explanation" and rule.get("defect_name") not in defect_names:
                    file.write('# ' + rule.get("defect_name") + '\n')
                    defect_names.add(rule.get("defect_name"))
                file.write(rule.get("rule_description") + '\n')
                if flag == "has_explanation":
                    file.write('# ' + rule.get("rule_explanation") + '\n')
                    file.write('\n')
    return {"url": folder_relative_path}
