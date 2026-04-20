from hsmolding.services import machine_service, polymer_service, project_service

from mdprocess.services import process_optimize_service

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import os
import re


# 以下属于模具参数
mold_mapping = [
    # 模具基本信息
    ("A5", "mold_no"),
    ("B5", "mold_type"),
    ("C5", "mold_name"),
    ("D5", "cavity_num"),
    ("E5", "inject_cycle_require"),
    ("F5", "subrule_no"),

    # 制品信息
    ("A9", "product_category"),
    ("B9", "product_type"),
    ("C9", "product_small_type"),
    ("D9", "product_name"),
    ("E9", "product_no"),
    ("F9", "product_total_weight"),
    ("G9", "product_projected_area"),

    # 冷却系统 - 型腔侧
    ("A18", "cavity_cooling_water_diameter"),
    ("B18", "cavity_cooling_circuit_number"),
    ("C18", "cavity_water_nozzle_specification"),

    # 冷却系统 - 型芯侧
    ("A21", "core_cooling_water_diameter"),
    ("B21", "core_cooling_circuit_number"),
    ("C21", "core_water_nozzle_specification"),

    # 顶出系统
    ("A25", "ejector_stroke"),
    ("B25", "ejector_rod_hole_diameter"),
    ("C25", "ejector_rod_hole_spacing"),
    ("D25", "ejector_rod_number"),
    ("E25", "ejector_force"),
    ("F25", "ejector_times"),
    ("G25", "reset_method"),
    ("H25", "ejection_method"),
    ("I25", "ejector_position_length"),
    ("J25", "ejector_position_width"),
    ("K25", "drain_distance"),

    # 物理属性与尺寸
    ("A29", "mold_weight"),
    ("B29", "hanging_mold_hole_specification"),
    ("A33", "size_horizon"),
    ("B33", "size_vertical"),
    ("C33", "size_thickness"),
    ("D33", "min_clamping_force"),
    ("E33", "mold_opening_stroke"),
    ("F33", "locate_ring_diameter"),

    # 责任人 & 时间
    ("A37", "customer"),
    ("B37", "project_engineer"),
    ("C37", "design_engineer"),
    ("D37", "production_engineer"),
    ("E37", "order_date"),
]
product_mapping = [
    # 第9+i行的数据
    ("H{row}", "ave_thickness"),
    ("I{row}", "max_thickness"),
    ("J{row}", "flow_length"),
    ("K{row}", "single_volume"),
    ("L{row}", "single_weight"),

    # 第13+i行的数据
    ("A{row2}", "sprue_hole_diameter"),
    ("B{row2}", "sprue_sphere_radius"),
    ("C{row2}", "runner_type"),
    ("D{row2}", "runner_length"),
    ("E{row2}", "runner_weight"),
    ("F{row2}", "gate_type"),
    ("G{row2}", "gate_num"),
    ("H{row2}", "gate_shape"),
    ("I{row2}", "gate_area"),
    ("J{row2}", "gate_radius"),
    ("K{row2}", "gate_length"),
    ("L{row2}", "gate_width"),
    ("M{row2}", "valve_num"),
    ("N{row2}", "hot_runner_num"),
]

# 以下属于材料参数
polymer_mapping = [
    # 基本信息
    ("A5", "abbreviation"),
    ("B5", "trademark"),
    ("C5", "manufacturer"),
    ("D5", "category"),
    ("E5", "data_source"),
    ("F5", "data_status"),
    ("G5", "internal_id"),
    ("H5", "level_code"),
    ("I5", "vendor_code"),

    # 温度 & 剪切参数
    ("A9", "max_melt_temperature"),
    ("B9", "min_melt_temperature"),
    ("C9", "recommend_melt_temperature"),
    ("D9", "max_mold_temperature"),
    ("E9", "min_mold_temperature"),
    ("F9", "recommend_mold_temperature"),
    ("G9", "max_shear_linear_speed"),
    ("H9", "min_shear_linear_speed"),
    ("I9", "recommend_shear_linear_speed"),
    ("J9", "degradation_temperature"),
    ("K9", "ejection_temperature"),
    ("L9", "recommend_injection_rate"),

    # 性能参数
    ("A11", "max_sheer_rate"),
    ("B11", "max_sheer_stress"),
    ("C11", "recommend_back_pressure"),
    ("D11", "barrel_residence_time"),

    # 干燥参数
    ("A14", "dry_method"),
    ("B14", "dry_temperature"),
    ("C14", "dry_time"),

    # WLF 模型参数
    ("A25", "cross_WLF_n"),
    ("B25", "cross_WLF_Tau"),
    ("C25", "cross_WLF_D1"),
    ("D25", "cross_WLF_D2"),
    ("E25", "cross_WLF_D3"),
    ("F25", "cross_WLF_A1"),
    ("G25", "cross_WLF_A2"),

    # 粘度模型参数
    ("A28", "c1"),
    ("B28", "c2"),
    ("C28", "switch_temp"),
    ("D28", "viscosity_index"),
    ("E28", "MFR_temp"),
    ("F28", "MFR_load"),
    ("G28", "MFR_measure"),

    # Tait PV-T 参数
    ("A18", "melt_density"),
    ("B18", "solid_density"),
    ("C18", "Tait_pvT_b5"),
    ("D18", "Tait_pvT_b6"),
    ("E18", "Tait_pvT_b1m"),
    ("F18", "Tait_pvT_b2m"),
    ("G18", "Tait_pvT_b3m"),
    ("H18", "Tait_pvT_b4m"),
    ("I18", "Tait_pvT_b1s"),
    ("J18", "Tait_pvT_b2s"),
    ("K18", "Tait_pvT_b3s"),
    ("L18", "Tait_pvT_b4s"),

    ("A20", "Tait_pvT_b7"),
    ("B20", "Tait_pvT_b8"),
    ("C20", "Tait_pvT_b9"),

    # 弹性模量与泊松比
    ("A33", "E1"),
    ("B33", "E2"),
    ("C33", "v12"),
    ("D33", "v23"),
    ("E33", "G12"),

    # 热膨胀系数
    ("A36", "Alpha1"),
    ("B36", "Alpha2"),

    # 收缩率
    ("A41", "average_horizontal_shrinkage"),
    ("B41", "average_vertical_shrinkage"),
    ("A44", "min_horizontal_shrinkage"),
    ("B44", "max_horizontal_shrinkage"),
    ("C44", "min_vertical_shrinkage"),
    ("D44", "max_vertical_shrinkage"),

    # 填料信息
    ("A49", "filler"),
    ("B49", "filler_type"),
    ("C49", "filler_shape"),
    ("D49", "filler_percentage"),
    ("E49", "filler_density"),
    ("F49", "filler_specific_heat"),
    ("G49", "filler_specific_thermal_conductivity"),

    # 填料力学性能
    ("A52", "filler_E1"),
    ("B52", "filler_E2"),
    ("C52", "filler_v12"),
    ("D52", "filler_v23"),
    ("E52", "filler_G12"),

    # 填料热膨胀与强度
    ("A55", "filler_Alpha1"),
    ("B55", "filler_Alpha2"),
    ("A59", "filler_horizontal_tensile_strength"),
    ("B59", "filler_vertical_tensile_strength"),
    ("C59", "filler_aspect_ratio"),
]

# 以下属于注塑机参数
machine_mapping = [
    # 基本信息
    ("A5", "data_source"),
    ("B5", "manufacturer"),
    ("C5", "trademark"),

    ("D5", "asset_no"),
    ("E5", "serial_no"),
    ("F5", "internal_id"),
    ("G5", "communication_interface"),
    ("H5", "agreement"),
    ("I5", "machine_type"),
    ("J5", "power_method"),
    ("K5", "propulsion_axis"),

    # 单位设置
    ("A9", "pressure_unit"),
    ("B9", "velocity_unit"),
    ("C9", "temperature_unit"),
    ("D9", "time_unit"),
    ("E9", "position_unit"),
    ("F9", "clamping_force_unit"),
    ("G9", "screw_rotation_unit"),
    ("H9", "power_unit"),
    ("I9", "backpressure_unit"),
    ("J9", "oc_pressure_unit"),
    ("K9", "oc_velocity_unit"),

    # 模具尺寸相关
    ("A30", "min_mold_size_horizon"),
    ("B30", "min_mold_size_vertical"),
    ("C30", "max_mold_size_horizon"),
    ("D30", "max_mold_size_vertical"),
    ("E30", "min_mold_thickness"),
    ("F30", "max_mold_thickness"),
    ("G30", "min_platen_opening"),
    ("H30", "max_platen_opening"),
    ("I30", "locate_ring_diameter"),

    # 拉杆信息
    ("A33", "pull_rod_size"),
    ("B33", "pull_rod_diameter"),
    ("C33", "pull_rod_distance_horizon"),
    ("D33", "pull_rod_distance_vertical"),

    # 锁模相关信息
    ("A36", "clamping_method"),
    ("B36", "max_clamping_force"),
    ("C36", "max_mold_open_stroke"),

    # 顶出信息
    ("A39", "max_ejection_force"),
    ("B39", "max_ejection_stroke"),
    ("C39", "ejection_hole_num"),

    # 系统参数
    ("A43", "hydraulic_system_pressure"),
    ("B43", "motor_power"),
    ("C43", "heater_power"),
    ("D43", "temp_control_zone_num"),
    ("E43", "needle_core"),
    ("F43", "core_pulling"),

    # 物理属性
    ("A47", "machine_weight"),
    ("B47", "size_length"),
    ("C47", "size_width"),
    ("D47", "size_height"),
    ("E47", "response_time"),
    ("F47", "enhancement_ratio"),
    ("G47", "manufacturing_date"),
    ("H47", "manufacture_date"),
    ("I47", "manufacture_no"),
    ("J47", "remark"),
]
injectors_info_mapping = [
    # 喷嘴信息
    ("B12", "serial_no"),
    
    ("A14", "nozzle_type"),
    ("B14", "nozzle_protrusion"),
    ("C14", "nozzle_hole_diameter"),
    ("D14", "nozzle_sphere_diameter"),
    ("E14", "nozzle_force"),

    # 螺杆信息
    ("A17", "screw_type"),
    ("B17", "screw_diameter"),
    ("C17", "screw_length_diameter_ratio"),
    ("D17", "screw_compression_ratio"),
    ("E17", "plasticizing_capacity"),
    ("F17", "barrel_heating_power"),
    ("G17", "max_injection_volume"),
    ("H17", "max_injection_weight"),
    ("I17", "max_injection_stroke"),

    # 性能参数
    ("A20", "max_injection_pressure"),
    ("B20", "max_injection_velocity"),
    ("C20", "max_holding_pressure"),
    ("D20", "max_holding_velocity"),
    ("E20", "max_metering_pressure"),
    ("F20", "max_screw_rotation_speed"),
    ("G20", "max_metering_back_pressure"),
    ("H20", "max_decompression_pressure"),
    ("I20", "max_decompression_velocity"),
    ("J20", "max_ejector_forward_velocity"),
    ("K20", "max_ejector_backward_velocity"),
    ("L20", "max_mold_opening_velocity"),
    ("M20", "max_mold_clamping_velocity"),

    # 设置性能参数
    ("A23", "max_set_injection_pressure"),
    ("B23", "max_set_injection_velocity"),
    ("C23", "max_set_holding_pressure"),
    ("D23", "max_set_holding_velocity"),
    ("E23", "max_set_metering_pressure"),
    ("F23", "max_set_screw_rotation_speed"),
    ("G23", "max_set_metering_back_pressure"),
    ("H23", "max_set_decompression_pressure"),
    ("I23", "max_set_decompression_velocity"),
    ("J23", "max_set_ejector_forward_velocity"),
    ("K23", "max_set_ejector_backward_velocity"),
    ("L23", "max_set_mold_opening_velocity"),
    ("M23", "max_set_mold_clamping_velocity"),

    # 阶段限制
    ("A26", "max_injection_stage"),
    ("B26", "max_holding_stage"),
    ("C26", "max_metering_stage"),
    ("D26", "max_temperature_stage"),
    ("E26", "max_opening_and_clamping_stage"),
    ("F26", "max_ejector_stage"),
]

# 以下属于工艺参数
# 配置：单元格位置 -> precondition 字段
precondition_mapping_report = [
    ("A4", "machine_data_source"),
    ("B4", "machine_trademark"),
    # ("C4", "serial_no"),
    ("C4", "polymer_abbreviation"),
    ("D4", "polymer_trademark"),
    ("E4", "recommend_melt_temperature"),
    ("A8", "mold_no"),
    ("B8", "cavity_num"),
    ("C8", "inject_cycle_require"),
    ("D8", "subrule_no"),
    ("A12", "runner_length"),
    ("B12", "runner_weight"),
    ("C12", "gate_type"),
    ("D12", "gate_num"),
    ("E12", "gate_shape"),
    ("F12", "gate_radius"),
    ("G12", "gate_length"),
    ("H12", "gate_width"),
    ("I12", "runner_type"),
    ("J12", "hot_runner_num"), 
    ("A16", "inject_part"),
    ("B16", "machine_serial_no"),
    ("D16", "product_no"),
    ("C16", "product_type"),
    ("E16", "product_name"),
    ("I16", "product_total_weight"),
    ("F16", "product_ave_thickness"),
    ("G16", "product_max_thickness"),
    ("H16", "product_max_length"),
]
process_config_report = [
    # 基础字段赋值：path 表示在 process_detail 中的路径
    {"type": "field", "cell": "A5", "path": ["inject_para", "injection_stage"]},
    {"type": "field", "cell": "A11", "path": ["inject_para", "injection_time"]},
    {"type": "field", "cell": "B11", "path": ["inject_para", "injection_delay_time"]},
    {"type": "field", "cell": "C11", "path": "inject_para.cooling_time"},
    {"type": "field", "cell": "A18", "path": ["holding_para", "holding_stage"]},
    {"type": "field", "cell": "A24", "path": ["metering_para", "metering_stage"]},
    {"type": "field", "cell": "A31", "path": ["metering_para", "decompressure_mode_before_metering"]},
    {"type": "field", "cell": "B31", "path": ["metering_para", "decompressure_mode_after_metering"]},
    {"type": "field", "cell": "A39", "path": ["metering_para", "metering_delay_time"]},
    {"type": "field", "cell": "B39", "path": ["metering_para", "metering_ending_position"]},
    {"type": "field", "cell": "A42", "path": ["temp_para", "barrel_temperature_stage"]},

    # 特殊字段：VP_switch
    {"type": "dict", "prefix": "VP_switch.", "items": {
        "A15": "VP_switch_mode",
        "B15": "VP_switch_time",
        "C15": "VP_switch_position",
        "D15": "VP_switch_pressure",
        "E15": "VP_switch_velocity"
    }},

    # 表格型数据：动态写入 sheet.cell(row, col)
    {"type": "table", "section_key": "inject_para.table_data", "start_row": 6, "start_col": 2},
    {"type": "table", "section_key": "holding_para.table_data", "start_row": 19, "start_col": 2},
    {"type": "table", "section_key": "metering_para.table_data", "start_row": 25, "start_col": 2},
    {"type": "table", "section_key": "temp_para.table_data", "start_row": 43, "start_col": 2},

    # 解压参数：特殊列表项
    {"type": "list", "section_key": "metering_para.decompressure_paras", "mapping": {
        0: {"cells": ("B35", "C35", "D35", "E35")},
        1: {"cells": ("B36", "C36", "D36", "E36")}
    }, "fields": ["pressure", "velocity", "distance", "time"]}
]
process_auxiliary_report = [
    {"type": "field", "cell": "A51", "path": ["mold_temp", "mold_temp_num"]},

    {"type": "table", "section_key": "hot_runner_temperatures", "start_row": 47, "start_col": 1},
    {"type": "table", "section_key": "mold_temp.mold_temp_list", "start_row": 53, "start_col": 1},
]

process_feedback_report = [
    {"type": "field", "cell": "A57", "path": "actual_product_weight"},
    {"type": "field", "cell": "A59", "path": "optimize_export.defect_name"},
    {"type": "field", "cell": "B59", "path": ["optimize_export", "defect_level"]},
    {"type": "field", "cell": "C59", "path": ["optimize_export", "defect_position"]},
    {"type": "field", "cell": "D59", "path": ["optimize_export", "rule_library_in_use"]},
    {"type": "field", "cell": "F59", "path": ["optimize_export", "rule_in_use"]},
    {"type": "field", "cell": "J59", "path": ["optimize_export", "defect_feedback"]},  
]

# 配置：单元格位置 -> machine_dict 中的键名
machine_mapping_report = [
    ("H6", "pressure_unit"),
    ("H7", "velocity_unit"),
    ("H8", "position_unit"),
    ("G19", "pressure_unit"),
    ("G20", "velocity_unit"),
    ("G21", "time_unit"),
    ("F25", "pressure_unit"),
    ("F26", "screw_rotation_unit"),
    ("F27", "pressure_unit"),
    ("F28", "position_unit"),
    ("B34", "pressure_unit"),
    ("C34", "velocity_unit"),
    ("D34", "position_unit"),
    ("E34", "time_unit"),    
]


from deepdiff import DeepDiff

def get_diff_dict(old_dict, new_dict):
    diff = DeepDiff(old_dict, new_dict, ignore_order=False, view='tree')
    changes = {}

    if 'values_changed' in diff:
        for item in diff['values_changed']:
            # 获取字段路径字符串，例如: root['process_detail'][0]['inject_para']['injection_stage']
            path = item.path()
            old_val = item.t1
            new_val = item.t2
            if isinstance(old_val, dict):
                # 检查 sections 列表中的变化
                old_sections = old_val.get('sections', [])
                new_sections = new_val.get('sections', [])
                for i in range(min(len(old_sections), len(new_sections))):
                    old_val = old_sections[i]
                    new_val = new_sections[i]
                    # 判断是否是从 0.0 变成非 0 的数值
                    if isinstance(old_val, (int, float)) and old_val in [0, 0.0] and isinstance(new_val, (int, float)) and new_val != 0:
                        changes[f"{path}['sections'][{i}]"] = (old_val, new_val)
            else:
                changes[path] = (old_val, new_val)

    return changes


def compare_neighbors(optimize_list):
    all_changes = []
    all_descriptions = []
    for i in range(len(optimize_list) - 1):
        old_item = optimize_list[i]
        new_item = optimize_list[i + 1]
        changes = get_diff_dict(old_item, new_item)
        all_changes.append(changes)
        descriptions = translate_diff_to_chinese(changes)
        all_descriptions.append(descriptions)
    return all_descriptions


def translate_diff_to_chinese(diff_dict):
    descriptions = []
    # 参数类型与字段名称的映射表
    PARAM_FIELD_MAP = {
        # ('参数类型', table_index) : "中文字段名"
        ("inject_para", 0): "注射压力",
        ("inject_para", 1): "注射速度",
        ("inject_para", 2): "注射位置",

        ("holding_para", 0): "保压压力",
        ("holding_para", 1): "保压时间",
        ("holding_para", 2): "保压位置",

        ("metering_para", 0): "计量压力",
        ("metering_para", 1): "计量螺杆转速",
        ("metering_para", 2): "计量背压",
        ("metering_para", 3): "计量位置",

        ("decompressure_paras", 0): "压力",
        ("decompressure_paras", 1): "速度",
        ("decompressure_paras", 2): "距离",
        ("decompressure_paras", 3): "时间",
    }
    SPECIAL_FIELD_MAP = {
        "VP_switch_mode": "切换方式",
        "VP_switch_time": "切换时间",
        "VP_switch_pressure": "切换压力",
        "VP_switch_velocity": "切换速度",
        "VP_switch_position": "VP切换位置",
        "decompressure_mode_before_metering": "储前松退模式",  
        "decompressure_mode_after_metering": "储后松退模式",  
        "injection_time": "注射时间",
        "injection_delay_time": "注射延迟",  
        "cooling_time": "冷却时间",
        "metering_delay_time": "储料延迟", 
        "metering_ending_position": "计量终止位置",  
    }
    for path, (old_val, new_val) in diff_dict.items():
        matched = False

        # 匹配特殊字段
        for key, chinese_name in SPECIAL_FIELD_MAP.items():
            if key in path:
                descriptions.append(f"{chinese_name}从 {old_val} → {new_val}")
                matched = True
                break

        if matched:
            continue
        # if "table_data" in path:
        try:
            # 提取 'inject_para' / 'holding_para' / 'metering_para'
            if "inject_para" in path:
                prefix_key = "inject_para"
            elif "holding_para" in path:
                prefix_key = "holding_para"
            elif "metering_para" in path:
                prefix_key = "metering_para"
            # elif "'cooling_para'" in path:
            #     prefix_key = "cooling_para"
            else:
                continue  # 不属于目标参数类型

            # 使用正则提取 table_data 索引
            pattern = r"\[(\d+)\]"
            table_match = re.findall(pattern, path)
            if not table_match:
                continue

            table_index = int(table_match[0])

            # 查找映射表
            key = (prefix_key, table_index)
            if key not in PARAM_FIELD_MAP:
                continue
            field_name = PARAM_FIELD_MAP[key]

            if len(table_match) >= 2:
                section_index = int(table_match[1])
                # 构造中文描述
                description = f"{field_name}第{section_index + 1}段从 {old_val} → {new_val}"
                descriptions.append(description)
        except Exception as e:
            # 路径格式不正确，跳过
            print(e)
            continue

        if "VP_switch_position" in path:
            descriptions.append(f"VP切换位置从 {old_val} → {new_val}")

        if "metering_ending_position" in path:
            descriptions.append(f"计量终止位置从 {old_val} → {new_val}")

        # elif "defect_info" in path and "level" in path:
        #     descriptions.append(f"反馈信息 - 缺陷等级从 '{old_val}' → '{new_val}'")

        # elif "defect_info" in path and "position" in path:
        #     descriptions.append(f"反馈信息 - 缺陷位置从 '{old_val}' → '{new_val}'")

        # elif "defect_info" in path and "count" in path:
        #     descriptions.append(f"反馈信息 - 缺陷数量从 {old_val} → {new_val}")
    return descriptions

mappings = {
    "SHORTSHOT": "短射",
    "SHRINKAGE": "缩水",
    "FLASH": "飞边",
    "GASVEINS": "气纹",
    "WELDLINE": "熔接痕",
    "MATERIALFLOWER": "料花",
    "AIRTRAP": "困气",
    "ABERRATION": "色差",
    "BURN": "烧焦",
    "WATERRIPPLE": "水波纹",
    "HARDDEMOLDING": "脱模不良",
    "TOPWHITE": "顶白",
    "WARPING": "变形",
    "OVERSIZE": "尺寸偏大",
    "UNDERSIZE": "尺寸偏小",
    "GATEMARK": "浇口印",
    "SHADING": "阴阳面"
}


# def wrap_text_by_chars(cell_value, max_chars=12):
#     """
#     将文本按指定的最大字符数分割，并插入换行符。
#     确保标点符号不会出现在下一行的开头。

#     :param cell_value: 单元格的内容
#     :param max_chars: 每行的最大字符数，默认为 12
#     :return: 换行后的字符串
#     """
#     if not cell_value:
#         return ""  # 如果内容为空，返回空字符串

#     # 定义标点符号集合（可以根据需要扩展）
#     punctuation = "，。！？；：,!?;:"

#     wrapped_lines = []
#     start = 0
#     while start < len(cell_value):
#         # 如果剩余字符数小于等于 max_chars，直接添加到结果中
#         if len(cell_value) - start <= max_chars:
#             wrapped_lines.append(cell_value[start:])
#             break

#         # 找到当前段落的分割点
#         end = start + max_chars

#         # 检查分割点是否在标点符号后
#         if cell_value[end] in punctuation:
#             # 如果分割点正好是标点符号，将标点符号留在当前行
#             wrapped_lines.append(cell_value[start : end + 1])
#             start = end + 1
#         elif cell_value[end - 1] in punctuation:
#             # 如果分割点前一个字符是标点符号，将标点符号留在当前行
#             wrapped_lines.append(cell_value[start:end])
#             start = end
#         else:
#             # 否则，找到最后一个空格或标点符号进行分割
#             split_pos = max(cell_value.rfind(" ", start, end), cell_value.rfind(punctuation, start, end))
#             if split_pos > start:
#                 # 如果找到了合适的分割点，使用该位置
#                 wrapped_lines.append(cell_value[start : split_pos + 1].rstrip())
#                 start = split_pos + 1
#             else:
#                 # 如果没有找到合适的分割点，强制分割
#                 wrapped_lines.append(cell_value[start:end])
#                 start = end

#     # 使用换行符连接分割后的行
#     return "\n".join(wrapped_lines)


def auto_adjust_row_height(ws, cell, font_size=16, padding=2, current_height=26):
    """
    根据单元格中的文字内容自动调整行高。

    :param ws: 工作表对象
    :param font_size: 字体大小，默认为11（Excel默认字体大小）
    :param padding: 行高额外增加的像素值，用于留白
    """
    if not cell.value:
        return 26
    # cell.value = wrap_text_by_chars(str(cell.value), max_chars=22)

    # 计算换行符分割后的行数
    lines = str(cell.value).split("\n")
    line_count = len(lines)

    # 近似计算所需行高
    # 假设每个字符宽度为字体大小的0.7倍，行高为字体大小的1.2倍
    estimated_height = line_count * font_size * 1.5 + padding

    height_result = 26
    if estimated_height and current_height:
        height_result = max(estimated_height, current_height)
    # 设置该行的行高
    ws.row_dimensions[cell.row].height = height_result
    cell.alignment = Alignment(wrap_text=True)
    return height_result


# 导出优化记录
def export_optimize(process_index_id):
    wb = load_workbook(settings.TEMPLATE_PATH + "standard/teplt_optimize.xlsx")
    optimization = process_optimize_service.get_process_optimization(process_index_id)
    precondition = optimization.get("precondition")
    optimize_list = optimization.get("optimize_list")
    sheet = wb["测试过程记录"]
    sheet["B1"] = optimization.get("created_at").strftime("%Y年%m月%d日")
    # B2 测试地点
    # B3 模具
    sheet["B3"] = precondition.get("mold_no")
    # B4 机器
    sheet["B4"] = precondition.get("machine_trademark")
    # B5 材料
    sheet["B5"] = precondition.get("polymer_trademark")
    changes = compare_neighbors(optimize_list)
    for num in range(len(optimize_list)):
        defect_desc = ""
        actual_product_weight = optimize_list[num].get("feedback_detail").get("actual_product_weight")
        defect_name = optimize_list[num].get("feedback_detail").get("optimize_export").get("defect_name")
        defect_level = optimize_list[num].get("feedback_detail").get("optimize_export").get("defect_level")
        if defect_name:
            defect_desc += f"缺陷{mappings.get(defect_name)},程度{defect_level}"
        if actual_product_weight:
            defect_desc += f"制品实际重量{actual_product_weight}g"
        sheet[f"B{9+num}"] = defect_desc
        if num < len(optimize_list) - 1:
            temp = "\n".join(changes[num])
            sheet[f"C{10+num}"] = temp
            if len(temp) >= 35:
                auto_adjust_row_height(sheet, sheet[f"C{10+num}"], font_size=11, padding=0, current_height=26)

    machine = None
    if precondition.get("machine_id"):
        machine: dict = machine_service.get_machine(precondition.get("machine_id"))
    sheet = wb["机器信息"]
    injectors_info: dict = machine.get("injectors_info")[0]
    apply_mapping(sheet, machine_mapping, machine)
    apply_mapping(sheet, injectors_info_mapping, injectors_info)

    sheet = wb["模具信息"]
    mold = project_service.get_mold_dict_by_id(precondition.get("mold_id"))
    product_infos = mold.get("product_infos")
    apply_mapping(sheet, mold_mapping, mold)
    if product_infos and len(product_infos) >=1:
        for i in range(len(product_infos)):
            product_info = product_infos[i]
            apply_mapping_multi(sheet, product_mapping, product_info, row=9 + i, row2=13 + i)
    if mold.get("assisting_equipments"):
        equipments = mold.get("assisting_equipments").split("|")
        for col in range(10):
            if sheet.cell(32, 7+col).value in equipments:
                sheet.cell(33, 7+col).value = "是"

    sheet = wb["材料信息"]
    polymer = polymer_service.get_polymer(precondition.get("polymer_id"))
    apply_mapping(sheet, polymer_mapping, polymer)

    sheet = wb["基本信息"]
    apply_mapping(sheet, precondition_mapping_report, precondition)

    # 首模
    if optimize_list:
        result_list = ["首模工艺"] + [f"opt#{i+1}" for i in range(len(optimize_list) - 1)]
        for num in range(len(optimize_list)-1):
            new_sheet = wb.copy_worksheet(wb["首模工艺"])
            new_sheet.title = f"opt#{num+1}"    
        for num in range(len(optimize_list)):
            process_record = optimize_list[num]
            sheet = wb[result_list[num]]
            apply_mapping(sheet, machine_mapping_report, machine)
            export_report_process(sheet, process_record, process_config_report, "process_detail")
            export_report_process(sheet, process_record, process_auxiliary_report, "auxiliary_detail")
            export_report_process(sheet, process_record, process_feedback_report, "feedback_detail")


    date = time.strftime("%Y%m%d%H%M%S", time.localtime())
    folder_relative_path = f"gsid_{mold.get('company_id')}/temp/"
    file_name = f"process_optimize_{mold.get('mold_no')}_{date}.xlsx"
    file_relative_path = f"{folder_relative_path}{file_name}"
    file_absolute_path = f"{settings.FILE_STORAGE_PATH}{file_relative_path}"

    check_path_existed(f"{settings.FILE_STORAGE_PATH}{folder_relative_path}")
    wb.save(file_absolute_path)

    return { "url": file_relative_path }


def check_path_existed(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


def apply_mapping(sheet, mapping, data):
    for cell, key in mapping:
        sheet[cell] = data.get(key)


def apply_mapping_multi(sheet, mapping, data, **kwargs):
    for cell_template, key in mapping:
        cell = cell_template.format(**kwargs)
        sheet[cell] = data.get(key)


def export_report_process(sheet, process_record, process_config_report, detail_type):
    if process_record and process_record.get(detail_type):
        process_detail = process_record.get(detail_type)
        print("*"*100, process_detail)
        apply_config(sheet, process_config_report, process_detail)


def apply_config(sheet, config, data):
    for item in config:
        if item["type"] == "field":
            cell = item["cell"]
            path = item["path"]

            if isinstance(path, str):
                keys = path.split(".")
            else:
                keys = path
            value = data
            try:
                for key in keys:
                    if isinstance(key, int) or (isinstance(key, str) and key.isdigit()):
                        key = int(key)
                    value = value[key]
            except (KeyError, IndexError, TypeError):
                value = None

            sheet[cell] = value

        elif item["type"] == "dict":
            prefix = item.get("prefix", "")
            items = item["items"]

            section = data
            for part in prefix.rstrip(".").split("."):
                if part in section:
                    section = section[part]
                else:
                    section = {}

            for cell, key in items.items():
                sheet[cell] = section.get(key)

        elif item["type"] == "table":
            section_path = item["section_key"]
            start_row = item["start_row"]
            start_col = item["start_col"]

            parts = section_path.split(".")
            table_data = data
            try:
                for part in parts:
                    if part.startswith("[") and part.endswith("]"):
                        index = int(part[1:-1])
                        table_data = table_data[index]
                    else:
                        table_data = table_data[part]
            except (KeyError, IndexError, TypeError):
                continue
            table_data = [] if not table_data else table_data
            for row_idx, row_data in enumerate(table_data):
                if row_data:
                    if isinstance(row_data, float):
                        sheet.cell(start_row, start_col + row_idx).value = row_data
                    elif isinstance(row_data, dict):
                        sections = row_data.get("sections", [])
                        for col_idx, val in enumerate(sections):
                            sheet.cell(start_row + row_idx, start_col + col_idx).value = val
                    elif isinstance(row_data, list):
                        sections = row_data[0].get("sections", [])
                        for col_idx, val in enumerate(sections):
                            sheet.cell(start_row + row_idx, start_col + col_idx).value = val

        elif item["type"] == "list":
            section_path = item["section_key"]
            mapping = item["mapping"]
            fields = item["fields"]

            parts = section_path.split(".")
            list_data = data
            try:
                for part in parts:
                    if part.startswith("[") and part.endswith("]"):
                        index = int(part[1:-1])
                        list_data = list_data[index]
                    else:
                        list_data = list_data[part]
            except (KeyError, IndexError, TypeError):
                continue

            for idx, info in enumerate(list_data):
                if idx not in mapping:
                    continue
                cells = mapping[idx]["cells"]
                for i, field in enumerate(fields):
                    sheet[cells[i]] = info.get(field)
